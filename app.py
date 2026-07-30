from random import randint
import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
from flask import Flask, render_template, request, Response, redirect, send_file, session, url_for, jsonify
from flask_login import LoginManager, login_required, UserMixin, current_user, login_user, logout_user
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy
import os
import platform
import cv2
import numpy as np
import csv
import face_recognition
from datetime import datetime, timedelta
from urllib.parse import parse_qs, urlencode, quote
import timeit
import time
import traceback
import pickle
import hashlib
import threading
from playsound import playsound
import pandas as pd
import json
import smtplib
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR       = os.path.join(BASE_DIR, 'static')
RECORDS_CSV_PATH = os.path.join(STATIC_DIR, 'records.csv')
TODAY_CSV_PATH   = os.path.join(STATIC_DIR, 'todayAttendance.csv')
HELP_JSON_PATH   = os.path.join(STATIC_DIR, 'help.json')
CAMERA_SOUND_PATH = os.path.join(STATIC_DIR, 'cameraSound.wav')
os.makedirs(STATIC_DIR, exist_ok=True)
app = Flask(__name__)
try:
    _np_major = int(np.__version__.split('.')[0])
    if _np_major >= 2:
        print("=" * 70)
        print(f"[CẢNH BÁO QUAN TRỌNG] Đang dùng numpy {np.__version__}.")
        print("dlib/face_recognition build sẵn thường KHÔNG tương thích với")
        print("numpy >= 2.0, gây lỗi 'Unsupported image type, must be 8bit")
        print("gray or RGB image' với MỌI ảnh (dù ảnh hoàn toàn hợp lệ).")
        print("CÁCH SỬA: mở terminal trong virtualenv của dự án, chạy:")
        print('    pip install "numpy<2" --force-reinstall')
        print("rồi khởi động lại server.")
        print("=" * 70)
except Exception:
    pass
app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///EmployeeDB.db"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JSON_AS_ASCII'] = False
try:
    app.json.ensure_ascii = False
except AttributeError:
    pass
app.config['SECRET_KEY'] = 'mysecretkey'
db = SQLAlchemy(app)
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com').strip()
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = (os.environ.get('MAIL_USERNAME') or 'f770654@gmail.com').strip()
app.config['MAIL_PASSWORD'] = (os.environ.get('MAIL_PASSWORD') or 'facerecogManisha').replace(' ', '').strip()
app.config['MAIL_DEFAULT_SENDER'] = app.config['MAIL_USERNAME']
TRANSLATIONS = {
    'vi': {
        'app_title': 'Hệ thống điểm danh nhân viên bằng nhận diện khuôn mặt',
        'nav_home': 'Trang chủ', 'nav_logout': 'Đăng xuất', 'nav_login': 'Đăng nhập',
        'nav_sessions': 'Buổi điểm danh', 'nav_attendance_sheet': 'Bảng điểm danh',
        'nav_recognizer': 'Nhận diện', 'nav_stats': 'Thống kê', 'nav_help': 'Trợ giúp',
        'nav_add_employee': 'Thêm nhân viên', 'nav_reports': 'Báo cáo tổng hợp',
        'greeting': 'Xin chào',
        'btn_add_employee': 'Thêm nhân viên mới', 'btn_recognizer': 'Nhận diện khuôn mặt',
        'btn_attendance_sheet': 'Bảng điểm danh', 'btn_sessions': 'Buổi điểm danh',
        'btn_stats': 'Thống kê', 'btn_help': 'Trợ giúp',

        'add_employee_title': 'Thêm nhân viên mới',
        'employee_id': 'Mã nhân viên', 'full_name': 'Họ và tên', 'department': 'Phòng ban',
        'email': 'Email', 'upload_photo_btn': 'Tải ảnh lên (rõ nét, không chỉnh sửa để có kết quả tốt nhất)',
        'modal_choice_title': 'Bạn muốn tải ảnh bằng cách nào?',
        'upload_from_files': 'Tải từ máy tính', 'take_photo': 'Chụp ảnh',
        'switch_to_upload': 'Chuyển sang: Tải ảnh lên', 'switch_to_camera': 'Chuyển sang: Chụp ảnh',
        'submit': 'Lưu thông tin', 'reset': 'Làm lại',
        'employee_database': 'Danh sách nhân viên', 'no_records': 'Chưa có dữ liệu',
        'table_sno': 'STT', 'table_id': 'Mã NV', 'table_name': 'Họ tên', 'table_dept': 'Phòng ban',
        'table_email': 'Email', 'table_hiring_date': 'Ngày vào làm', 'table_photo': 'Ảnh', 'table_action': 'Thao tác',
        'delete': 'Xoá', 'update': 'Cập nhật', 'cancel': 'Huỷ',
        'confirm_delete_suffix': 'sẽ bị xoá vĩnh viễn khỏi hệ thống',
        'warning_no_photo': 'Lưu ý: bạn chưa tải ảnh lên. Nên thêm ảnh để hệ thống có thể tự động điểm danh cho nhân viên này.',
        'error_duplicate_id': 'Mã nhân viên hoặc tên đăng nhập đã tồn tại. Vui lòng thử lại.',

        'start_recognition': 'Bắt đầu nhận diện', 'back_to_home': 'Quay về trang chủ',
        'status_not_started': 'Chưa bắt đầu.', 'status_recognizing': 'Đang nhận diện...',
        'auto_record_note': 'Khi hiện khung xanh, hệ thống tự động ghi nhận điểm danh ngay lập tức — không cần chờ thao tác nào thêm.',

        'lang_switch': 'EN',
    },
    'en': {
        'app_title': 'Face Recognition Employee Attendance System',
        'nav_home': 'Home', 'nav_logout': 'Logout', 'nav_login': 'Login',
        'nav_sessions': 'Attendance Sessions', 'nav_attendance_sheet': 'Attendance Sheet',
        'nav_recognizer': 'Recognizer', 'nav_stats': 'Statistics', 'nav_help': 'Help',
        'nav_add_employee': 'Add Employee', 'nav_reports': 'Reports Dashboard',
        'greeting': 'Hey',
        'btn_add_employee': 'Add New Employee', 'btn_recognizer': 'Recognizer',
        'btn_attendance_sheet': 'Attendance Sheet', 'btn_sessions': 'Attendance Sessions',
        'btn_stats': 'Statistics', 'btn_help': 'Help',

        'add_employee_title': 'Add New Employee',
        'employee_id': 'Employee ID', 'full_name': 'Full Name', 'department': 'Department',
        'email': 'Email', 'upload_photo_btn': 'Upload Photo (clear & unfiltered for best results)',
        'modal_choice_title': 'How do you want to upload the photo?',
        'upload_from_files': 'Upload from files', 'take_photo': 'Take photo',
        'switch_to_upload': 'Switch to: Upload photo', 'switch_to_camera': 'Switch to: Take photo',
        'submit': 'Submit', 'reset': 'Reset',
        'employee_database': 'Employee Database', 'no_records': 'No Records',
        'table_sno': 'S.No', 'table_id': 'ID', 'table_name': 'Name', 'table_dept': 'Dept',
        'table_email': 'Email', 'table_hiring_date': 'Hiring Date', 'table_photo': 'Photo', 'table_action': 'Action',
        'delete': 'Delete', 'update': 'Update', 'cancel': 'Cancel',
        'confirm_delete_suffix': 'will be permanently deleted from the system',
        'warning_no_photo': "Warning: you did not upload a photo. Add one so this employee's attendance can be recorded automatically.",
        'error_duplicate_id': 'Employee with the same ID or username already exists. Please try again.',

        'start_recognition': 'Start recognition', 'back_to_home': 'Back to home',
        'status_not_started': 'Not started.', 'status_recognizing': 'Recognizing...',
        'auto_record_note': 'As soon as a green box appears, attendance is recorded automatically — no further action needed.',

        'lang_switch': 'VI',
    },
}
@app.context_processor
def inject_i18n():
    lang = session.get('lang', 'vi')
    if lang not in TRANSLATIONS:
        lang = 'vi'

    def t(key):
        return TRANSLATIONS[lang].get(key, TRANSLATIONS['vi'].get(key, key))
    return dict(t=t, current_lang=lang)
@app.route('/set_lang/<lang>')
def set_lang(lang):
    if lang in TRANSLATIONS:
        session['lang'] = lang
    return redirect(request.referrer or '/')
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
mail_ = Mail(app)
encodedList = []
imgNames    = []
cap         = None
cap2        = None
pic         = None
bot_responses = {}
last_video_error   = None 
last_camera_status = "Chưa mở"
last_frame_shape    = None
MATCH_THRESHOLD = 0.50
RECOGNITION_SEMAPHORE = threading.Semaphore(1)
SEMAPHORE_WAIT_TIMEOUT = 20
MAX_LIVE_FRAME_DIM = 480
def _downscale_for_recognition(img):
    h, w = img.shape[:2]
    m = max(h, w)
    if m <= MAX_LIVE_FRAME_DIM:
        return img
    scale = MAX_LIVE_FRAME_DIM / m
    return cv2.resize(img, (int(w * scale), int(h * scale)))
def count_faces_lightweight(img):
    acquired = RECOGNITION_SEMAPHORE.acquire(timeout=SEMAPHORE_WAIT_TIMEOUT)
    if not acquired:
        return 0, "Server đang bận xử lý yêu cầu khác, vui lòng thử lại sau vài giây."
    try:
        small = _downscale_for_recognition(img)
        try:
            img_rgb = to_rgb(small)
            locs = face_recognition.face_locations(img_rgb, number_of_times_to_upsample=1, model="hog")
            return len(locs), None
        except Exception as e:
            return 0, f"{type(e).__name__}: {e}"
    finally:
        RECOGNITION_SEMAPHORE.release()
ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.webp'}
ALLOWED_IMAGE_MIMETYPES = {'image/jpeg', 'image/png', 'image/bmp', 'image/webp'}
def _is_allowed_image_file(file_storage):
    if not file_storage or not file_storage.filename:
        return False
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        return False
    mimetype = (file_storage.mimetype or '').lower()
    if mimetype and mimetype not in ALLOWED_IMAGE_MIMETYPES:
        return False
    return True
def _validate_and_save_photo(file_storage, dest_path):
    if not _is_allowed_image_file(file_storage):
        return False, "Chỉ chấp nhận file ảnh (jpg, jpeg, png, bmp, webp)."
    file_bytes = file_storage.read()
    npbuf = np.frombuffer(file_bytes, dtype=np.uint8)
    img = cv2.imdecode(npbuf, cv2.IMREAD_COLOR)
    if img is None:
        return False, "Không đọc được file ảnh. Vui lòng chọn file ảnh hợp lệ."
    n_faces, err = count_faces_lightweight(img)
    if err:
        return False, err
    if n_faces == 0:
        return False, "Không phát hiện khuôn mặt trong ảnh. Vui lòng chọn ảnh khác."
    cv2.imwrite(dest_path, img)
    return True, None
def _validate_and_save_captured(img, dest_path):
    if img is None:
        return False, "Không nhận được ảnh chụp."
    n_faces, err = count_faces_lightweight(img)
    if err:
        return False, err
    if n_faces == 0:
        return False, "Không phát hiện khuôn mặt trong ảnh chụp. Vui lòng chụp lại."
    cv2.imwrite(dest_path, img)
    return True, None
def _decode_image_from_request():
    import base64
    img = None
    for field_name in ('frame', 'image', 'file', 'photo'):
        if field_name in request.files:
            file_bytes = request.files[field_name].read()
            npbuf = np.frombuffer(file_bytes, dtype=np.uint8)
            img = cv2.imdecode(npbuf, cv2.IMREAD_COLOR)
            break
    if img is None and request.is_json:
        data = request.get_json(silent=True) or {}
        b64_str = data.get('image') or data.get('frame') or data.get('file')
        if b64_str:
            if ',' in b64_str and b64_str.strip().startswith('data:'):
                b64_str = b64_str.split(',', 1)[1]
            try:
                file_bytes = base64.b64decode(b64_str)
                npbuf = np.frombuffer(file_bytes, dtype=np.uint8)
                img = cv2.imdecode(npbuf, cv2.IMREAD_COLOR)
            except Exception as e:
                return None, f"Không giải mã được base64: {e}"

    if img is None:
        return None, ("Không tìm thấy ảnh trong request. Cần multipart field "
                       "'frame'/'image'/'file'/'photo', hoặc JSON "
                       "{'image': 'data:...base64...'}.")
    return img, None
def _test_frame_readable(c, tries=5):
    for _ in range(tries):
        ok, frame = c.read()
        if ok and frame is not None and frame.size > 0:
            if float(np.std(frame)) > 1.0:
                return True, frame.shape
        time.sleep(0.15)
    return False, None
def open_camera(index=0, width=None, height=None):
    global last_camera_status
    candidates = []
    if platform.system() == "Windows":
        candidates = [
            (index, cv2.CAP_DSHOW),
            (index, cv2.CAP_MSMF),
            (index, cv2.CAP_ANY),
        ]
    else:
        candidates = [(index, cv2.CAP_ANY)]
    extra_indexes = [i for i in (0, 1, 2) if i != index]
    for i in extra_indexes:
        if platform.system() == "Windows":
            candidates += [(i, cv2.CAP_DSHOW), (i, cv2.CAP_MSMF)]
        else:
            candidates += [(i, cv2.CAP_ANY)]
    for idx, backend in candidates:
        try:
            c = cv2.VideoCapture(idx, backend)
        except Exception as e:
            print(f"[CAMERA] Lỗi khi mở index={idx} backend={backend}: {e}")
            continue
        if not c.isOpened():
            c.release()
            continue
        if width is not None:
            c.set(cv2.CAP_PROP_FRAME_WIDTH, width)
        if height is not None:
            c.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
        ok, shape = _test_frame_readable(c)
        if ok:
            msg = f"OK (index={idx}, backend={backend}, frame={shape})"
            print(f"[CAMERA] Mở thành công: {msg}")
            last_camera_status = msg
            return c
        else:
            print(f"[CAMERA] index={idx} backend={backend} isOpened()=True "
                  "nhưng KHÔNG đọc được frame hợp lệ (toàn đen/timeout). Thử tiếp...")
            c.release()
    last_camera_status = "THẤT BẠI: không mở được camera nào, hoặc mở được nhưng không đọc được frame hợp lệ"
    print(f"[CAMERA ERROR] {last_camera_status}. "
          "Kiểm tra: (1) camera đang bị app khác (Zoom, Teams, app Camera "
          "của Windows...) chiếm dụng — tắt hết các app đó rồi thử lại; "
          "(2) Settings > Privacy & Security > Camera đã bật quyền cho "
          "trình duyệt/Python; (3) driver camera; (4) thử đổi index.")
    return cv2.VideoCapture()
def to_rgb(img):
    channels = img.shape[2] if img.ndim == 3 else 1
    if channels == 4:
        rgb = cv2.cvtColor(img, cv2.COLOR_BGRA2RGB)
    elif channels == 1:
        rgb = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)
    else:
        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    return np.ascontiguousarray(rgb, dtype=np.uint8)
def ensure_records_csv():
    csv_path = RECORDS_CSV_PATH
    if not os.path.exists(csv_path):
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            f.write(','.join(RECORDS_FIELDS) + '\n')
        return
    _repair_records_csv()
RECORDS_FIELDS = ['Id', 'Name', 'Department', 'Time', 'Date', 'Trạng thái']
def _repair_records_csv():
    csv_path = RECORDS_CSV_PATH
    try:
        with open(csv_path, newline='', encoding='utf-8-sig') as f:
            raw_rows = list(csv.reader(f))
    except Exception:
        return
    if not raw_rows:
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            f.write(','.join(RECORDS_FIELDS) + '\n')
        return
    fixed_rows = []
    for row in raw_rows[1:]:
        row = [c.strip() for c in row]
        if len(row) == 0:
            continue
        if len(row) == len(RECORDS_FIELDS):
            fixed_rows.append(row)
        elif len(row) > len(RECORDS_FIELDS):
            head = row[:1]
            tail = row[-4:]
            name_dept = row[1:-4]
            merged_name = ' '.join(name_dept[:max(1, len(name_dept) - 1)]) if name_dept else ''
            merged_dept = name_dept[-1] if len(name_dept) > 1 else (name_dept[0] if name_dept else '')
            fixed_rows.append(head + [merged_name, merged_dept] + tail)
        else:
            padded = row + [''] * (len(RECORDS_FIELDS) - len(row))
            fixed_rows.append(padded)
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(RECORDS_FIELDS)
        writer.writerows(fixed_rows)
def _read_records_df():
    ensure_records_csv()
    try:
        return pd.read_csv(RECORDS_CSV_PATH, engine='python', on_bad_lines='skip', encoding='utf-8-sig')
    except TypeError:
        return pd.read_csv(RECORDS_CSV_PATH, engine='python', error_bad_lines=False, encoding='utf-8-sig')
@login_manager.user_loader
def load_user(user_id):
    return users.query.get(user_id)
class employee(db.Model):
    id          = db.Column(db.String(20), primary_key=True)
    name        = db.Column(db.String(20), nullable=False)
    department  = db.Column(db.String(20), nullable=False)
    email       = db.Column(db.String(20), nullable=False)
    role        = db.Column(db.String(20), nullable=False, default='student')
    hiringDate  = db.Column(db.String(10), default=lambda: datetime.now().strftime("%d-%m-%Y"))
    def __repr__(self):
        return f"{self.id} - {self.name} - {self.department} - {self.email} - {self.hiringDate}"
class users(db.Model, UserMixin):
    id          = db.Column(db.String(20), primary_key=True)
    username    = db.Column(db.String(20), nullable=False, unique=True)
    name        = db.Column(db.String(80), nullable=True)
    mail        = db.Column(db.String(80), nullable=True)
    password    = db.Column(db.String(80), nullable=False)
    role        = db.Column(db.String(20), nullable=False, default='admin')
    status      = db.Column(db.String(20), nullable=False, default='active')
    workplace   = db.Column(db.String(80), nullable=True)
    position    = db.Column(db.String(80), nullable=True)
    dateCreated = db.Column(db.DateTime, default=datetime.utcnow)
    def __repr__(self):
        return '<User {}>'.format(self.username)
class Subject(db.Model):
    id   = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
class Class_(db.Model):
    __tablename__ = 'class_'
    id    = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name  = db.Column(db.String(50), nullable=False, unique=True)
    grade = db.Column(db.String(10), nullable=True)
class TeacherAssignment(db.Model):
    id         = db.Column(db.Integer, primary_key=True, autoincrement=True)
    teacher_id = db.Column(db.String(20), nullable=False)
    class_name = db.Column(db.String(50), nullable=False)
class EditHistory(db.Model):
    id          = db.Column(db.Integer, primary_key=True, autoincrement=True)
    admin_id    = db.Column(db.String(20), nullable=False)
    target_id   = db.Column(db.String(20), nullable=False)
    action      = db.Column(db.String(200), nullable=False)
    timestamp   = db.Column(db.DateTime, default=datetime.utcnow)
class AttendanceSession(db.Model):
    id           = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name         = db.Column(db.String(50), nullable=False)
    days_of_week = db.Column(db.String(40), nullable=False)
    open_time    = db.Column(db.String(5), nullable=False)
    close_time   = db.Column(db.String(5), nullable=False)
    def day_list(self):
        return [d for d in self.days_of_week.split(',') if d]
    def __repr__(self):
        return f"{self.id} - {self.name} - {self.days_of_week} - {self.open_time}-{self.close_time}"
ROLE_LABELS = {'student': 'Học sinh', 'teacher': 'Giáo viên'}
ROLE_DEPT_LABEL = {'student': 'Lớp', 'teacher': 'Môn học giảng dạy'}
ACCOUNT_STATUS_LABELS = {'pending': 'Chờ duyệt', 'active': 'Đang hoạt động', 'disabled': 'Đã vô hiệu hoá'}
path = os.path.join(STATIC_DIR, 'TrainingImages')
if not os.path.exists(path):
    os.makedirs(path)
PENDING_PHOTOS_DIR = os.path.join(STATIC_DIR, 'pending_photos')
os.makedirs(PENDING_PHOTOS_DIR, exist_ok=True)
from functools import wraps
def require_role(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect('/login')
            if getattr(current_user, 'role', None) not in roles:
                return jsonify({"error": "Bạn không có quyền truy cập chức năng này."}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator
def _log_edit(admin_id, target_id, action):
    db.session.add(EditHistory(admin_id=admin_id, target_id=target_id, action=action))
    db.session.commit()
def _allowed_employee_ids_for(u):
    if not u.is_authenticated:
        return set()
    if u.role == 'admin':
        return None
    if u.role == 'student':
        return {u.id}
    if u.role == 'teacher':
        assigned = {a.class_name for a in TeacherAssignment.query.filter_by(teacher_id=u.id).all()}
        ids = set()
        if assigned:
            ids = {e.id for e in employee.query.filter(employee.department.in_(assigned)).all()}
        ids.add(u.id)
        return ids
    return set()
@app.route('/')
def index():
    global cap, cap2
    for c in (cap, cap2):
        try:
            if c is not None:
                c.release()
        except Exception as e:
            print("[index] release lỗi:", e)
    return render_template('index.html')
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = users.query.filter_by(username=username).first()
        if user is not None and user.password == password:
            if user.status == 'pending':
                return render_template('login.html', incorrect=True,
                                       msg="Tài khoản của bạn đang chờ quản trị viên phê duyệt.")
            if user.status == 'disabled':
                return render_template('login.html', incorrect=True,
                                       msg="Tài khoản của bạn đã bị vô hiệu hoá. Vui lòng liên hệ quản trị viên.")
            login_user(user)
            return redirect('/')
        return render_template('login.html', incorrect=True)
    return render_template('login.html')
@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    logout_user()
    return redirect('/')
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        id        = request.form['id']
        username  = request.form['username']
        name      = request.form['name']
        workplace = request.form.get('workplace', '').strip()
        position  = request.form.get('position', '').strip()
        mail      = request.form['mail'].strip()
        pass1     = request.form['pass']
        pass2     = request.form['pass2']
        user  = users.query.filter_by(username=username).first()
        user2 = users.query.filter_by(id=id).first()
        if user is not None or user2 is not None:
            return render_template('signup.html', incorrect=True,
                                   msg='Mã số hoặc tên đăng nhập đã tồn tại.')
        if pass1 != pass2:
            return render_template('signup.html', incorrect=True, msg="Mật khẩu không khớp.")
        import re
        EMAIL_RE = r'^[^@\s]+@[^@\s]+\.[^@\s]+$'
        if not re.match(EMAIL_RE, mail):
            return render_template('signup.html', incorrect=True,
                                   msg="Email không hợp lệ. Vui lòng nhập đúng định dạng (VD: ten@vidu.com).")
        existing_mail = users.query.filter_by(mail=mail).first()
        if existing_mail is not None:
            return render_template('signup.html', incorrect=True,
                                   msg="Email này đã được đăng ký cho một tài khoản khác.")
        otp = randint(100000, 999999)
        ok, err = _sendResetMail(mail, otp)
        if not ok:
            if err and (err.startswith("smtp_auth:") or err.startswith("smtp_other:")):
                friendly_msg = err.split(":", 1)[1]
            else:
                friendly_msg = "Không gửi được email xác nhận. Vui lòng thử lại sau."
            return render_template('signup.html', incorrect=True, msg=friendly_msg)
        session['otp'] = otp
        session['pending_admin'] = dict(
            id=id, name=name, mail=mail, username=username, pass1=pass1,
            workplace=workplace, position=position,
        )
        return render_template('OTP.html')
    return render_template('signup.html')
@app.route("/registerEmployee", methods=['GET', 'POST'])
def registerEmployee():
    invalid = 0
    photo_error = None
    is_admin = current_user.is_authenticated and current_user.role == 'admin'
    if request.method == 'POST':
        if is_admin:
            role = request.form.get('role', 'student')
            if role not in ('teacher', 'student'):
                role = 'student'
            invalid, photo_error = _admin_add_employee(role)
            if not invalid:
                return redirect(f'/manage?active_tab={role}')
        else:
            invalid, photo_error, otp_sent = _process_employee_registration()
            if otp_sent:
                return render_template('OTP.html')
    default_role = request.args.get('role', 'student')
    if default_role not in ('teacher', 'student'):
        default_role = 'student'
    return render_template("registerEmployee.html", invalid=invalid, photo_error=photo_error,
                           classes=_sorted_classes(), subjects=Subject.query.all(),
                           is_admin=is_admin, default_role=default_role)
def _process_employee_registration(fixed_role=None):
    global pic
    id       = request.form['id']
    name     = request.form['name']
    dept     = request.form['dept']
    mail     = request.form['mail'].strip()
    role     = fixed_role or request.form.get('role', 'student')
    username = request.form.get('username', '').strip()
    pass1    = request.form.get('pass', '')
    pass2    = request.form.get('pass2', '')
    if role not in ('teacher', 'student'):
        role = 'student'
    if (employee.query.filter_by(id=id).first() is not None
            or users.query.filter_by(id=id).first() is not None
            or users.query.filter_by(username=username).first() is not None):
        return 1, None, False
    if not username or pass1 != pass2 or len(pass1) < 8:
        return 3, None, False
    tmp_photo_path = os.path.join(PENDING_PHOTOS_DIR, id + '.jpg')
    if pic is not None:
        ok_photo, photo_error = _validate_and_save_captured(pic, tmp_photo_path)
        pic = None
    else:
        photo = request.files.get('photo')
        if photo and photo.filename:
            ok_photo, photo_error = _validate_and_save_photo(photo, tmp_photo_path)
        else:
            return 2, None, False
    if not ok_photo:
        return 5, photo_error, False
    otp = randint(100000, 999999)
    ok, err = _sendResetMail(mail, otp)
    if not ok:
        return 4, None, False
    session['otp'] = otp
    session['pending_employee'] = dict(
        id=id, name=name, dept=dept, mail=mail, role=role,
        username=username, pass1=pass1, photo_tmp_path=tmp_photo_path,
    )
    return 0, None, True
def _admin_add_employee(role):
    global pic
    id       = request.form.get('id', '').strip()
    name     = request.form.get('name', '').strip()
    dept     = request.form.get('dept', '').strip()
    mail     = request.form.get('mail', '').strip()
    username = request.form.get('username', '').strip()
    pass1    = request.form.get('pass', '')
    pass2    = request.form.get('pass2', '')
    if role not in ('teacher', 'student'):
        role = 'student'
    if not (id and name and dept and mail and username):
        return 3, None
    if (employee.query.filter_by(id=id).first() is not None
            or users.query.filter_by(id=id).first() is not None
            or users.query.filter_by(username=username).first() is not None):
        return 1, None
    if pass1 != pass2 or len(pass1) < 8:
        return 3, None
    import re as _re
    if not _re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', mail):
        return 3, None
    dest_path = os.path.join(path, id + '.jpg')
    if pic is not None:
        ok_photo, photo_error = _validate_and_save_captured(pic, dest_path)
        pic = None
    else:
        photo = request.files.get('photo')
        if photo and photo.filename:
            ok_photo, photo_error = _validate_and_save_photo(photo, dest_path)
        else:
            return 2, None
    if not ok_photo:
        return 5, photo_error
    db.session.add(employee(id=id, name=name, department=dept, email=mail, role=role))
    db.session.add(users(id=id, name=name, mail=mail, username=username,
                         password=pass1, role=role, status='active'))
    db.session.commit()
    _log_edit(current_user.id, id, f"thêm {ROLE_LABELS.get(role, role)} mới qua trang quản lý")
    return 0, None
def _manage_redirect(qs, **overrides):
    params = {k: v[0] for k, v in parse_qs(qs or '').items()}
    for k, v in overrides.items():
        if v is None:
            params.pop(k, None)
        else:
            params[k] = v
    params = {k: v for k, v in params.items() if v not in (None, '')}
    return redirect(url_for('manage_page') + (('?' + urlencode(params)) if params else ''))
@app.route("/add", methods=['GET', 'POST'])
@login_required
@require_role('admin')
def add():
    return redirect('/registerEmployee')
@app.route("/manage/<string:role>", methods=['GET', 'POST'])
@login_required
@require_role('admin')
def manage_role(role):
    if role not in ('teacher', 'student'):
        role = 'student'
    return redirect(f'/manage?active_tab={role}')
@app.route('/manage', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def manage_page():
    global cap2, pic
    try:
        if cap2 is not None:
            cap2.release()
    except Exception as e:
        print("[manage_page] cap2.release() lỗi:", e)
    if request.method == 'POST':
        qs = request.form.get('qs', '')
        active_tab = request.form.get('active_tab', 'student')
        return _manage_redirect(qs, active_tab=active_tab)
    active_tab = request.args.get('active_tab', 'student').strip()
    if active_tab not in ('student', 'teacher', 'accounts', 'approve', 'classes', 'subjects', 'assignments'):
        active_tab = 'student'
    edit_uid = request.args.get('edit_uid', '').strip()
    invalid = session.pop('manage_invalid', 0)
    photo_error = session.pop('photo_error', None)
    q_student = request.args.get('q_student', '').strip()
    dept_student = request.args.get('dept_student', '').strip()
    q_teacher = request.args.get('q_teacher', '').strip()
    dept_teacher = request.args.get('dept_teacher', '').strip()
    q_accounts = request.args.get('q_accounts', '').strip()
    role_accounts = request.args.get('role_accounts', '').strip()
    status_accounts = request.args.get('status_accounts', '').strip()
    q_classes = request.args.get('q_classes', '').strip()
    q_subjects = request.args.get('q_subjects', '').strip()
    filter_teacher_assign = request.args.get('filter_teacher_assign', '').strip()
    filter_class_assign = request.args.get('filter_class_assign', '').strip()
    student_users = users.query.filter_by(role='student').order_by(users.name).all()
    student_emp_map = {e.id: e for e in employee.query.filter_by(role='student').all()}
    if q_student:
        ql = q_student.lower()
        student_users = [u for u in student_users if ql in u.id.lower() or ql in (u.name or '').lower()]
    if dept_student:
        student_users = [u for u in student_users
                          if student_emp_map.get(u.id) and student_emp_map[u.id].department == dept_student]
    teacher_users = users.query.filter_by(role='teacher').order_by(users.name).all()
    teacher_emp_map = {e.id: e for e in employee.query.filter_by(role='teacher').all()}
    if q_teacher:
        ql = q_teacher.lower()
        teacher_users = [u for u in teacher_users if ql in u.id.lower() or ql in (u.name or '').lower()]
    if dept_teacher:
        teacher_users = [u for u in teacher_users
                          if teacher_emp_map.get(u.id) and teacher_emp_map[u.id].department == dept_teacher]
    all_users = users.query.order_by(users.role, users.name).all()
    emp_map = {e.id: e for e in employee.query.all()}
    if q_accounts:
        ql = q_accounts.lower()
        all_users = [u for u in all_users if ql in u.id.lower() or ql in (u.name or '').lower()
                     or ql in (u.username or '').lower()]
    if role_accounts:
        all_users = [u for u in all_users if u.role == role_accounts]
    if status_accounts:
        all_users = [u for u in all_users if u.status == status_accounts]
    pending  = users.query.filter(users.role.in_(['teacher', 'student']), users.status == 'pending').all()
    active   = users.query.filter(users.role.in_(['teacher', 'student']), users.status == 'active').all()
    disabled = users.query.filter(users.role.in_(['teacher', 'student']), users.status == 'disabled').all()
    classes_all = _sorted_classes()
    if q_classes:
        ql = q_classes.lower()
        classes_all = [c for c in classes_all if ql in c.name.lower()]
    subjects_all = Subject.query.all()
    if q_subjects:
        ql = q_subjects.lower()
        subjects_all = [s for s in subjects_all if ql in s.name.lower()]
    teachers = employee.query.filter_by(role='teacher').all()
    assignments = TeacherAssignment.query.all()
    if filter_teacher_assign:
        assignments = [a for a in assignments if a.teacher_id == filter_teacher_assign]
    if filter_class_assign:
        assignments = [a for a in assignments if a.class_name == filter_class_assign]
    return render_template(
        'managePage.html', active_tab=active_tab, edit_uid=edit_uid, invalid=invalid, photo_error=photo_error,
        student_users=student_users, student_emp_map=student_emp_map,
        teacher_users=teacher_users, teacher_emp_map=teacher_emp_map,
        all_users=all_users, emp_map=emp_map,
        pending=pending, active=active, disabled=disabled,
        classes_all=classes_all, subjects_all=subjects_all,
        classes=_sorted_classes(), subjects=Subject.query.all(),
        teachers=teachers, assignments=assignments,
        role_labels=ROLE_LABELS, status_labels=ACCOUNT_STATUS_LABELS,
        q_student=q_student, dept_student=dept_student,
        q_teacher=q_teacher, dept_teacher=dept_teacher,
        q_accounts=q_accounts, role_accounts=role_accounts, status_accounts=status_accounts,
        q_classes=q_classes, q_subjects=q_subjects,
        filter_teacher_assign=filter_teacher_assign, filter_class_assign=filter_class_assign,
    )
def _build_xlsx(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'DanhSach'
    font_name = 'Times New Roman'
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name=font_name, size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26
    header_row = 3
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name=font_name, size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        cell.fill = header_fill
    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name=font_name, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
    for col_idx in range(1, n_cols + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in rows:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else ''
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 10), 40)
    ws.print_title_rows = f'{header_row}:{header_row}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.5)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
def _xlsx_response(title, headers, rows, filename):
    output = _build_xlsx(title, headers, rows)
    return Response(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={filename}'})
@app.route('/manage/export/<string:section>')
@login_required
@require_role('admin')
def manage_export(section):
    if section in ('student', 'teacher'):
        role_users = users.query.filter_by(role=section).order_by(users.name).all()
        emp_map = {e.id: e for e in employee.query.filter_by(role=section).all()}
        headers = ['Mã số', 'Họ tên', 'Vai trò', 'Lớp/Môn học', 'Email', 'Ngày vào', 'Trạng thái']
        rows = [[u.id, u.name, ROLE_LABELS.get(u.role, u.role),
                emp_map.get(u.id).department if emp_map.get(u.id) else '', u.mail,
                emp_map.get(u.id).hiringDate if emp_map.get(u.id) else '',
                ACCOUNT_STATUS_LABELS.get(u.status, u.status)] for u in role_users]
        title = 'DANH SÁCH HỌC SINH' if section == 'student' else 'DANH SÁCH GIÁO VIÊN'
        return _xlsx_response(title, headers, rows, f'danh_sach_{section}.xlsx')
    elif section == 'accounts':
        all_users = users.query.order_by(users.role, users.name).all()
        emp_map = {e.id: e for e in employee.query.all()}
        headers = ['Mã số', 'Tên đăng nhập', 'Họ tên', 'Vai trò', 'Trạng thái',
                   'Email', 'Lớp/Môn học', 'Nơi công tác', 'Chức vụ']
        rows = [[u.id, u.username, u.name, ROLE_LABELS.get(u.role, u.role),
                ACCOUNT_STATUS_LABELS.get(u.status, u.status), u.mail,
                emp_map.get(u.id).department if emp_map.get(u.id) else '',
                u.workplace or '', u.position or ''] for u in all_users]
        return _xlsx_response('DANH SÁCH TÀI KHOẢN', headers, rows, 'danh_sach_tai_khoan.xlsx')
    elif section == 'approve':
        recs = users.query.filter(users.role.in_(['teacher', 'student'])).order_by(users.status, users.name).all()
        headers = ['Mã số', 'Họ tên', 'Vai trò', 'Trạng thái', 'Email']
        rows = [[u.id, u.name, ROLE_LABELS.get(u.role, u.role),
                ACCOUNT_STATUS_LABELS.get(u.status, u.status), u.mail] for u in recs]
        return _xlsx_response('DUYỆT TÀI KHOẢN', headers, rows, 'duyet_tai_khoan.xlsx')
    elif section == 'classes':
        headers = ['Khối', 'Lớp']
        rows = [[c.grade or '', c.name] for c in _sorted_classes()]
        return _xlsx_response('DANH SÁCH LỚP HỌC', headers, rows, 'danh_sach_lop.xlsx')
    elif section == 'subjects':
        headers = ['Môn học']
        rows = [[s.name] for s in Subject.query.all()]
        return _xlsx_response('DANH SÁCH MÔN HỌC', headers, rows, 'danh_sach_mon_hoc.xlsx')
    elif section == 'assignments':
        teacher_map = {e.id: e.name for e in employee.query.filter_by(role='teacher').all()}
        headers = ['Giáo viên', 'Lớp']
        rows = [[teacher_map.get(a.teacher_id, a.teacher_id), a.class_name] for a in TeacherAssignment.query.all()]
        return _xlsx_response('PHÂN CÔNG GIÁO VIÊN', headers, rows, 'phan_cong_giao_vien.xlsx')
    else:
        return redirect('/manage')
def gen_frames_takePhoto():
    global cap2, pic
    start             = timeit.default_timer()
    flag              = False
    num               = -1
    last_countdown_t  = time.time()
    while True:
        if cap2 is None or not cap2.isOpened():
            print("[takePhoto] Camera cap2 không sẵn sàng, dừng stream.")
            break
        ret, frame = cap2.read()
        if not ret:
            time.sleep(0.05)
            continue
        frame = cv2.flip(frame, 1)
        small = _downscale_for_recognition(frame)
        h0, w0 = frame.shape[:2]
        h1, w1 = small.shape[:2]
        INV = (h0 / h1) if h1 else 1.0
        try:
            img_rgb = to_rgb(small)
            facesLoc = face_recognition.face_locations(img_rgb, number_of_times_to_upsample=1, model="hog")
        except Exception as e:
            print("[takePhoto] Lỗi dò khuôn mặt:", e)
            facesLoc = []
        if not facesLoc:
            flag = False
        else:
            for faceLoc in facesLoc:
                try:
                    from deepface import DeepFace
                    result = DeepFace.analyze(frame, actions=['emotion'],
                                              enforce_detection=False, silent=True)
                    if isinstance(result, list):
                        result = result[0]
                    dominant_emotion = result['dominant_emotion']
                except Exception as e:
                    print("[takePhoto] DeepFace lỗi:", e)
                    dominant_emotion = 'neutral'
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = int(y1 * INV), int(x2 * INV), int(y2 * INV), int(x1 * INV)
                elapsed = timeit.default_timer() - start
                if dominant_emotion == 'happy' and elapsed > 5:
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    if flag:
                        cv2.putText(frame, str(num),
                                    (150, 200), cv2.FONT_HERSHEY_SIMPLEX,
                                    6, (255, 255, 255), 20)
                        if time.time() - last_countdown_t >= 1.0:
                            num -= 1
                            last_countdown_t = time.time()
                    else:
                        flag            = True
                        num             = 3
                        last_countdown_t = time.time()
                else:
                    flag = False
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    if elapsed <= 5:
                        remaining = int(5 - elapsed) + 1
                        cv2.putText(frame, f"Get ready: {remaining}s",
                                    (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX,
                                    0.7, (0, 165, 255), 2)
        _, buffer = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n'
               + buffer.tobytes() + b'\r\n')
@app.route('/takePhoto', methods=['GET', 'POST'])
def takePhoto():
    global cap2
    try:
        if cap2 is not None:
            cap2.release()
    except Exception as e:
        print("[takePhoto route] release lỗi:", e)
    cap2 = open_camera(0)
    return Response(gen_frames_takePhoto(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')
@app.route('/detect_face_live', methods=['POST'])
def detect_face_live():
    img, err = _decode_image_from_request()
    if img is None:
        return jsonify({"face_detected": False, "message": err}), 400
    n_faces, face_err = count_faces_lightweight(img)
    if face_err:
        return jsonify({"face_detected": False, "message": face_err}), 500
    return jsonify({"face_detected": n_faces >= 1, "faces_detected": n_faces})
@app.route('/capture_photo', methods=['POST'])
def capture_photo():
    global pic
    img, err = _decode_image_from_request()
    if img is None:
        return jsonify({"success": False, "message": err}), 400
    faces_detected, face_err = count_faces_lightweight(img)
    if face_err:
        print(f"[capture_photo] Lỗi kiểm tra khuôn mặt: {face_err}")
        return jsonify({
            "success": False,
            "message": f"Lỗi xử lý ảnh (dlib/numpy): {face_err}",
            "faces_detected": 0,
        }), 500
    if faces_detected == 0:
        return jsonify({
            "success": False,
            "message": "Không phát hiện khuôn mặt nào trong ảnh. Vui lòng "
                       "chụp lại, đảm bảo đủ sáng và nhìn thẳng camera.",
            "faces_detected": 0,
        })
    if faces_detected > 1:
        return jsonify({
            "success": False,
            "message": "Phát hiện nhiều hơn 1 khuôn mặt. Chỉ chụp 1 người.",
            "faces_detected": faces_detected,
        })
    pic = img.copy()
    return jsonify({
        "success": True,
        "message": "Chụp ảnh thành công.",
        "faces_detected": faces_detected,
    })
@app.route('/check_face', methods=['POST'])
@login_required
@require_role('admin', 'teacher')
def check_face():
    img, err = _decode_image_from_request()
    if img is None:
        return jsonify({"faces_detected": 0, "ok": False, "message": err}), 400
    faces_detected, face_err = count_faces_lightweight(img)
    if face_err:
        return jsonify({"faces_detected": 0, "ok": False,
                        "message": f"Lỗi xử lý ảnh: {face_err}"}), 500
    if faces_detected == 0:
        return jsonify({"faces_detected": 0, "ok": False,
                        "message": "Chưa thấy khuôn mặt nào trong khung hình."})
    if faces_detected > 1:
        return jsonify({"faces_detected": faces_detected, "ok": False,
                        "message": "Phát hiện nhiều hơn 1 khuôn mặt — chỉ chụp/upload ảnh 1 người."})
    return jsonify({"faces_detected": 1, "ok": True, "message": "OK — đúng 1 khuôn mặt."})
ENCODING_CACHE_PATH = os.path.join(STATIC_DIR, 'encodings_cache.pkl')
def _training_folder_signature(myList):
    parts = []
    for fname in sorted(myList):
        try:
            fpath = os.path.join(path, fname)
            stat = os.stat(fpath)
            parts.append(f"{fname}:{stat.st_mtime_ns}:{stat.st_size}")
        except Exception:
            parts.append(fname)
    joined = "|".join(parts)
    return hashlib.md5(joined.encode('utf-8')).hexdigest()
def load_known_faces(force=False):
    global encodedList, imgNames
    _t0 = timeit.default_timer()

    try:
        if not os.path.exists(path):
            os.makedirs(path)
        myList = os.listdir(path)
    except Exception as e:
        print(f"[encode] LỖI nghiêm trọng: không đọc được thư mục '{path}': {e}")
        traceback.print_exc()
        myList = []
    myList = [f for f in myList if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    signature = _training_folder_signature(myList)
    if not force:
        try:
            with open(ENCODING_CACHE_PATH, 'rb') as f:
                cache = pickle.load(f)
            if cache.get('signature') == signature:
                encodedList = cache['encodedList']
                imgNames    = cache['imgNames']
                print(f"[encode] Dùng CACHE — bỏ qua tính lại "
                      f"({len(encodedList)} khuôn mặt, ảnh training không đổi). "
                      f"Mất {timeit.default_timer() - _t0:.3f}s.")
                return
        except (FileNotFoundError, EOFError, KeyError, pickle.PickleError):
            pass
        except Exception as e:
            print(f"[encode] Không đọc được cache (bỏ qua, tính lại): {e}")
    images   = []
    imgNames = []
    for fname in myList:
        img_path = os.path.join(path, fname)
        try:
            img = cv2.imread(img_path)
        except Exception as e:
            print(f"[encode] cv2.imread thất bại ({fname}): {e}")
            continue
        if img is None:
            print(f"[encode] Không đọc được ảnh (None): {img_path}")
            continue
        images.append(img)
        emp_id = os.path.splitext(fname)[0].split('_')[0]
        imgNames.append(emp_id)
    encodedList = []
    for i, img in enumerate(images):
        try:
            h, w = img.shape[:2]
            max_dim = 800
            if max(h, w) > max_dim:
                scale = max_dim / max(h, w)
                img = cv2.resize(img, (int(w * scale), int(h * scale)))
            img_rgb = to_rgb(img)
            locs = face_recognition.face_locations(img_rgb, number_of_times_to_upsample=2)
            if not locs:
                print(f"[encode] Không tìm thấy khuôn mặt trong: {imgNames[i]}")
                continue
            encodings = face_recognition.face_encodings(img_rgb, locs, num_jitters=2)
            if encodings:
                encodedList.append(encodings[0])
            else:
                print(f"[encode] face_encodings trả rỗng: {imgNames[i]}")
        except Exception as e:
            print(f"[encode] Lỗi khi xử lý {imgNames[i]}: {e}")
    print(f"[encode] Đã nạp thành công {len(encodedList)}/{len(images)} khuôn mặt "
          f"(tính mới, đã lưu cache). Mất {timeit.default_timer() - _t0:.3f}s.")
    try:
        with open(ENCODING_CACHE_PATH, 'wb') as f:
            pickle.dump({
                'signature': signature,
                'encodedList': encodedList,
                'imgNames': imgNames,
            }, f)
    except Exception as e:
        print(f"[encode] Không lưu được cache (không ảnh hưởng hoạt động, chỉ mất tốc độ lần sau): {e}")
@app.route("/encode")
@login_required
@require_role('admin', 'teacher')
def encode():
    try:
        load_known_faces()
    except Exception as e:
        print(f"[encode route] Lỗi không lường trước: {e}")
        traceback.print_exc()
    return render_template("recogPage.html")
def mark_attendance_entry(emp_id):
    ensure_records_csv()
    with app.app_context():
        emp = employee.query.filter_by(id=emp_id).first()
    if emp is None:
        print(f"[chấm công] CẢNH BÁO: nhận diện được ID '{emp_id}' nhưng "
              f"không tìm thấy hồ sơ nhân viên tương ứng trong database "
              f"— có thể hồ sơ đã bị xoá nhưng ảnh training còn sót lại.")
        return False
    now   = datetime.now()
    dtime = now.strftime('%H:%M:%S')
    date  = now.strftime('%d-%m-%Y')
    minute_key = dtime[:5]
    with open(RECORDS_CSV_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get('Id') == emp_id and row.get('Date') == date and row.get('Time', '')[:5] == minute_key:
                print(f"[chấm công] Bỏ qua: {emp_id} đã được ghi nhận trong phút {minute_key} ngày {date}.")
                return False
    with open(RECORDS_CSV_PATH, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow([emp_id, emp.name, emp.department, dtime, date, 'Có mặt'])
        f.flush()
        os.fsync(f.fileno())
    print(f"[chấm công] Đã ghi vào {RECORDS_CSV_PATH}: "
          f"{emp_id} - {emp.name} lúc {dtime} ngày {date}")
    return True
def recognize_and_annotate(img, draw=True, draw_hud=False):
    global encodedList, imgNames, last_video_error
    faces_info = []
    error_message = None
    facesCurFrame   = []
    encodesCurFrame = []
    acquired = RECOGNITION_SEMAPHORE.acquire(timeout=SEMAPHORE_WAIT_TIMEOUT)
    if not acquired:
        return [], "Server đang bận xử lý yêu cầu khác, vui lòng thử lại sau vài giây."
    try:
        h_orig, w_orig = img.shape[:2]
        detect_img = _downscale_for_recognition(img)
        h_det, w_det = detect_img.shape[:2]
        scale_back = (w_orig / w_det) if w_det else 1.0
        try:
            img_rgb       = to_rgb(detect_img)
            facesCurFrame = face_recognition.face_locations(
                                img_rgb, number_of_times_to_upsample=1, model="hog")
            encodesCurFrame = face_recognition.face_encodings(img_rgb, facesCurFrame, num_jitters=1)
            last_video_error = None
        except Exception as e:
            error_message = f"{type(e).__name__}: {e}"
            last_video_error = error_message
            print(f"[recognize] LỖI face_recognition: {error_message}")
            traceback.print_exc()
            facesCurFrame   = []
            encodesCurFrame = []
    finally:
        RECOGNITION_SEMAPHORE.release()
    h_img = img.shape[0]
    facesCurFrame = [
        tuple(int(round(v * scale_back)) for v in loc) for loc in facesCurFrame
    ]
    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        y1, x2, y2, x1 = faceLoc
        y1, x2, y2, x1 = int(y1), int(x2), int(y2), int(x1)
        try:
            if len(encodedList) > 0:
                faceDis = face_recognition.face_distance(encodedList, encodeFace)
                per_id_best = {}
                for dist, name in zip(faceDis, imgNames):
                    if name not in per_id_best or dist < per_id_best[name]:
                        per_id_best[name] = dist
                ranked = sorted(per_id_best.items(), key=lambda kv: kv[1])
                Id_candidate, best_dist = ranked[0]
                MIN_MARGIN = 0.08
                second_dist = ranked[1][1] if len(ranked) > 1 else None
                ambiguous = (second_dist is not None
                             and (second_dist - best_dist) < MIN_MARGIN
                             and best_dist < MATCH_THRESHOLD)
                if best_dist < MATCH_THRESHOLD and not ambiguous:
                    Id = Id_candidate
                    try:
                        with app.app_context():
                            emp = employee.query.filter_by(id=Id).first()
                        name = emp.name if emp else Id
                    except Exception:
                        name = Id
                    conf_pct = int((1 - best_dist) * 100)
                    if draw:
                        text_y1  = min(y2 + 25, h_img - 50)
                        text_y2  = min(y2 + 50, h_img - 25)
                        text_y3  = min(y2 + 75, h_img - 5)
                        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                        cv2.putText(img, f"ID: {Id}",
                                    (x1, text_y1), cv2.FONT_HERSHEY_TRIPLEX,
                                    0.7, (0, 255, 0), 2)
                        cv2.putText(img, name,
                                    (x1, text_y2), cv2.FONT_HERSHEY_TRIPLEX,
                                    0.7, (0, 255, 0), 2)
                        cv2.putText(img, f"Conf: {conf_pct}%",
                                    (x1, text_y3), cv2.FONT_HERSHEY_SIMPLEX,
                                    0.5, (0, 255, 0), 1)
                    try:
                        marked = mark_attendance_entry(Id)
                        if marked:
                            print(f"[recognize] Nhận diện & chấm công: {Id} - {name} (dist={best_dist:.3f})")
                    except Exception as e:
                        print(f"[recognize] mark_attendance_entry lỗi: {e}")
                        traceback.print_exc()
                    faces_info.append({
                        "box": [x1, y1, x2, y2], "status": "known",
                        "id": Id, "name": name, "confidence": conf_pct,
                    })
                else:
                    if draw:
                        text_y = min(y2 + 25, h_img - 10)
                        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 0, 255), 2)
                        label = 'unclear (2 nguoi giong nhau)' if ambiguous else 'unknown'
                        cv2.putText(img, label,
                                    (x1, text_y), cv2.FONT_HERSHEY_TRIPLEX,
                                    0.6, (0, 0, 255), 2)
                    faces_info.append({
                        "box": [x1, y1, x2, y2],
                        "status": "ambiguous" if ambiguous else "unknown",
                        "id": None, "name": None, "confidence": None,
                    })
            else:
                if draw:
                    text_y = min(y2 + 25, h_img - 10)
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    cv2.putText(img, 'Please encode first',
                                (x1, text_y), cv2.FONT_HERSHEY_SIMPLEX,
                                0.6, (0, 0, 255), 1)
                faces_info.append({
                    "box": [x1, y1, x2, y2], "status": "no_encoding",
                    "id": None, "name": None, "confidence": None,
                })
        except Exception as e:
            print(f"[recognize] LỖI khi vẽ khung nhận diện: {e}")
            traceback.print_exc()
    if draw and draw_hud:
        hud_lines = [
            f"Cam: {last_camera_status}",
            f"Faces detected: {len(facesCurFrame)}  |  Encoded DB: {len(encodedList)}",
        ]
        if last_video_error:
            hud_lines.append(f"Last error: {last_video_error[:70]}")
        for i, line in enumerate(hud_lines):
            y = 55 + i * 22
            cv2.rectangle(img, (5, y - 16), (min(630, 15 + len(line) * 9), y + 4), (0, 0, 0), -1)
            cv2.putText(img, line, (10, y), cv2.FONT_HERSHEY_SIMPLEX,
                        0.5, (0, 255, 255), 1)
    return faces_info, error_message
def gen_frames():
    global cap, last_video_error
    frame_count      = 0
    consecutive_fail = 0
    MAX_CONSEC_FAIL  = 60
    if cap is None or not cap.isOpened():
        err_img = np.zeros((480, 640, 3), dtype=np.uint8)
        msg1 = "KHONG MO DUOC CAMERA"
        msg2 = (last_camera_status or "")[:60]
        msg3 = "Kiem tra: camera dang bi app khac dung, quyen Camera trong"
        msg4 = "Windows Settings, hoac vao /diagnostics de xem chi tiet."
        cv2.putText(err_img, msg1, (30, 200), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
        cv2.putText(err_img, msg2, (30, 240), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 200, 255), 1)
        cv2.putText(err_img, msg3, (30, 270), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        cv2.putText(err_img, msg4, (30, 295), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (200, 200, 200), 1)
        _, buffer = cv2.imencode('.jpg', err_img)
        frame_bytes = (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n'
                       + buffer.tobytes() + b'\r\n')
        for _ in range(50):
            yield frame_bytes
            time.sleep(0.2)
        print("[video] Camera cap không sẵn sàng ngay từ đầu, dừng stream.")
        return
    while True:
        if cap is None or not cap.isOpened():
            print("[video] Camera cap không sẵn sàng, dừng stream.")
            last_video_error = "Camera không mở (cap is None hoặc isOpened()=False)"
            break
        success, img = cap.read()
        if not success or img is None or img.size == 0:
            consecutive_fail += 1
            last_video_error = f"cap.read() thất bại (lần liên tiếp: {consecutive_fail})"
            if consecutive_fail >= MAX_CONSEC_FAIL:
                print(f"[video] {last_video_error}. Vượt ngưỡng, dừng stream.")
                break
            time.sleep(0.05)
            continue
        consecutive_fail = 0
        frame_count += 1
        img = cv2.flip(img, 1)
        cv2.putText(img, datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    (10, 25), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (0, 0, 255), 1)
        recognize_and_annotate(img, draw=True, draw_hud=True)
        if frame_count % 60 == 0:
            print(f"[video] frame={frame_count} encodedDB={len(encodedList)}")
        _, buffer = cv2.imencode('.jpg', img)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n'
               + buffer.tobytes() + b'\r\n')
@app.route('/video', methods=['GET', 'POST'])
@login_required
@require_role('admin', 'teacher')
def video():
    global cap
    try:
        if cap is not None:
            cap.release()
    except Exception as e:
        print("[video route] release lỗi:", e)
    try:
        load_known_faces()
    except Exception as e:
        print(f"[video route] load_known_faces() lỗi không lường trước: {e}")
        traceback.print_exc()
    try:
        cap = open_camera(0, width=640, height=480)
    except Exception as e:
        print(f"[video route] open_camera() lỗi không lường trước: {e}")
        traceback.print_exc()
        cap = cv2.VideoCapture() 
    return Response(gen_frames(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')
@app.route('/stop_video')
@login_required
def stop_video():
    global cap
    try:
        if cap is not None:
            cap.release()
    except Exception as e:
        print("[stop_video] release lỗi:", e)
    return redirect('/')
@app.route('/recognize_frame', methods=['POST'])
def recognize_frame():
    import base64
    img = None
    for field_name in ('frame', 'image', 'file'):
        if field_name in request.files:
            file_bytes = request.files[field_name].read()
            npbuf = np.frombuffer(file_bytes, dtype=np.uint8)
            img = cv2.imdecode(npbuf, cv2.IMREAD_COLOR)
            break
    if img is None and request.is_json:
        data = request.get_json(silent=True) or {}
        b64_str = data.get('image') or data.get('frame') or data.get('file')
        if b64_str:
            if ',' in b64_str and b64_str.strip().startswith('data:'):
                b64_str = b64_str.split(',', 1)[1]
            try:
                file_bytes = base64.b64decode(b64_str)
                npbuf = np.frombuffer(file_bytes, dtype=np.uint8)
                img = cv2.imdecode(npbuf, cv2.IMREAD_COLOR)
            except Exception as e:
                return jsonify({"error": f"Không giải mã được base64: {e}"}), 400
    if img is None:
        return jsonify({
            "error": "Không tìm thấy ảnh trong request. Cần multipart field "
                     "'frame'/'image'/'file', hoặc JSON {'image': 'data:...base64...'}."
        }), 400
    try:
        faces_info, error_message = recognize_and_annotate(img, draw=False)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500
    first_known = next((f for f in faces_info if f["status"] == "known"), None)
    return jsonify({
        "recognized": first_known is not None,
        "name": first_known["name"] if first_known else None,
        "id": first_known["id"] if first_known else None,
        "faces": faces_info,
        "encoded_count": len(encodedList),
        "error": error_message,
    })
@app.route('/diagnostics')
@login_required
@require_role('admin')
def diagnostics():
    report = {}
    try:
        import dlib
        report['dlib_version'] = getattr(dlib, '__version__', 'không rõ version')
        report['dlib_import_ok'] = True
    except Exception as e:
        report['dlib_import_ok'] = False
        report['dlib_import_error'] = f"{type(e).__name__}: {e}"
    report['opencv_version'] = cv2.__version__
    report['numpy_version']  = np.__version__
    try:
        report['numpy_dlib_warning'] = (
            int(np.__version__.split('.')[0]) >= 2
        ) 
    except Exception:
        report['numpy_dlib_warning'] = None
    try:
        report['face_recognition_import_ok'] = True
    except Exception as e:
        report['face_recognition_import_ok'] = False
    test_cap = open_camera(0, width=640, height=480)
    report['camera_status'] = last_camera_status
    report['camera_isOpened'] = bool(test_cap.isOpened())
    try:
        test_cap.release()
    except Exception:
        pass
    training_report = []
    for fname in os.listdir(path):
        if not fname.lower().endswith(('.png', '.jpg', '.jpeg')):
            continue
        img_path = os.path.join(path, fname)
        entry = {'file': fname}
        try:
            img = cv2.imread(img_path)
            if img is None:
                entry['status'] = 'KHÔNG ĐỌC ĐƯỢC FILE (None)'
            else:
                h, w = img.shape[:2]
                entry['size'] = f"{w}x{h}"
                img_rgb = to_rgb(img)
                locs = face_recognition.face_locations(img_rgb, number_of_times_to_upsample=2)
                entry['faces_found'] = len(locs)
                entry['status'] = 'OK' if locs else 'KHÔNG DÒ ĐƯỢC KHUÔN MẶT TRONG ẢNH NÀY'
        except Exception as e:
            entry['status'] = f"LỖI: {type(e).__name__}: {e}"
        training_report.append(entry)
    report['training_images'] = training_report
    report['encoded_count'] = len(encodedList)
    report['encoded_names'] = imgNames
    return jsonify(report)
@app.route("/delete/<string:id>")
@login_required
@require_role('admin')
def delete(id):
    emp = employee.query.filter_by(id=id).first()
    if emp is None:
        return redirect("/accounts")
    role_for_redirect = emp.role if emp.role in ('teacher', 'student') else 'student'
    db.session.delete(emp)
    linked_user = users.query.filter_by(id=id).first()
    if linked_user is not None and linked_user.role != 'admin':
        db.session.delete(linked_user)
    db.session.commit()
    _log_edit(current_user.id, id, "xoá tài khoản qua trang quản lý")
    try:
        for fname in os.listdir(path):
            stem = os.path.splitext(fname)[0]
            if stem.split('_')[0] == id:
                os.unlink(os.path.join(path, fname))
    except Exception as e:
        print("[delete] Không xoá được ảnh:", e)
    ensure_records_csv()
    df = _read_records_df()
    df.loc[df["Id"] == id, "Status"] = "Terminated"
    df.to_csv(RECORDS_CSV_PATH, index=False, encoding='utf-8-sig')
    return redirect(f"/manage?active_tab={role_for_redirect}")
@app.route("/update", methods=['GET', 'POST'])
@login_required
@require_role('admin')
def update():
    global pic
    id  = request.form['id']
    emp = employee.query.filter_by(id=id).first()
    if emp is None:
        return redirect("/accounts")
    emp.name       = request.form['name']
    emp.department = request.form['dept']
    emp.email      = request.form['mail']
    linked_user = users.query.filter_by(id=id).first()
    if linked_user is not None:
        linked_user.name = emp.name
        linked_user.mail = emp.email
    db.session.commit()
    fileNm = id + '.jpg'
    update_photo_error = None
    try:
        photo = request.files.get('photo')
        if photo and photo.filename:
            ok_photo, update_photo_error = _validate_and_save_photo(photo, os.path.join(path, fileNm))
        elif pic is not None:
            ok_photo, update_photo_error = _validate_and_save_captured(pic, os.path.join(path, fileNm))
            pic = None
    except Exception as e:
        print("[update] Lỗi cập nhật ảnh:", e)
    if update_photo_error:
        session['photo_error'] = update_photo_error
    _log_edit(current_user.id, id, "cập nhật thông tin qua trang quản lý")
    ensure_records_csv()
    df = _read_records_df()
    df.loc[df["Id"] == id, ['Name', 'Department']] = [emp.name, emp.department]
    df.to_csv(RECORDS_CSV_PATH, index=False, encoding='utf-8-sig')
    return redirect(f"/manage?active_tab={linked_user.role if linked_user else 'student'}")
@app.route('/approve')
@login_required
@require_role('admin')
def approve_list():
    return redirect('/manage?active_tab=approve')
@app.route('/approve/<string:uid>/<string:action>')
@login_required
@require_role('admin')
def approve_action(uid, action):
    u = users.query.filter_by(id=uid).first()
    if u is not None and u.role != 'admin':
        if action == 'activate':
            u.status = 'active'
            db.session.commit()
            _log_edit(current_user.id, uid, "duyệt tài khoản (kích hoạt)")
        elif action == 'disable':
            u.status = 'disabled'
            db.session.commit()
            _log_edit(current_user.id, uid, "vô hiệu hoá tài khoản")
    next_url = request.args.get('next')
    return redirect(next_url or '/manage?active_tab=approve')
@app.route('/accounts')
@login_required
@require_role('admin')
def accounts_list():
    return redirect('/manage?active_tab=accounts')
@app.route('/accounts/edit/<string:uid>', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def account_edit(uid):
    global pic
    u = users.query.filter_by(id=uid).first()
    if u is None:
        return redirect('/manage?active_tab=accounts')
    emp = employee.query.filter_by(id=uid).first()
    target_tab = u.role if u.role in ('teacher', 'student') else 'accounts'
    if request.method == 'GET':
        return redirect(f'/manage?active_tab={target_tab}&edit_uid={uid}')
    qs = request.form.get('qs', f'active_tab={target_tab}')
    invalid = 0
    photo_error = None
    new_username = request.form.get('username', '').strip()
    new_name     = request.form.get('name', '').strip()
    new_mail     = request.form.get('mail', '').strip()
    new_status   = request.form.get('status', u.status)
    new_dept     = request.form.get('dept', '').strip()
    new_pass     = request.form.get('new_pass', '')
    new_pass2    = request.form.get('new_pass2', '')
    dup = users.query.filter(users.username == new_username, users.id != uid).first()
    if dup is not None:
        invalid = 1
    elif new_pass and (new_pass != new_pass2 or len(new_pass) < 8):
        invalid = 3
    else:
        changes = []
        if u.username != new_username:
            changes.append(f"tên đăng nhập: {u.username} -> {new_username}")
        if u.name != new_name:
            changes.append(f"họ tên: {u.name} -> {new_name}")
        if u.mail != new_mail:
            changes.append(f"email: {u.mail} -> {new_mail}")
        if u.role != 'admin' and u.status != new_status:
            changes.append(f"trạng thái: {u.status} -> {new_status}")
        if new_pass:
            changes.append("đổi mật khẩu")
        u.username = new_username
        u.name     = new_name
        u.mail     = new_mail
        if u.role != 'admin':
            u.status = new_status
        if new_pass:
            u.password = new_pass
        if emp is not None:
            emp.name  = new_name
            emp.email = new_mail
            if new_dept and new_dept != emp.department:
                changes.append(f"{'môn học' if u.role == 'teacher' else 'lớp'}: {emp.department} -> {new_dept}")
                emp.department = new_dept
        photo = request.files.get('photo')
        fileNm = uid + '.jpg'
        if photo and photo.filename:
            ok_photo, photo_error = _validate_and_save_photo(photo, os.path.join(path, fileNm))
            if ok_photo:
                changes.append("cập nhật ảnh khuôn mặt")
            else:
                invalid = 5
        elif pic is not None:
            ok_photo, photo_error = _validate_and_save_captured(pic, os.path.join(path, fileNm))
            pic = None
            if ok_photo:
                changes.append("cập nhật ảnh khuôn mặt (chụp)")
            else:
                invalid = 5
        db.session.commit()
        if changes:
            _log_edit(current_user.id, uid, "; ".join(changes))
        ensure_records_csv()
        df = _read_records_df()
        if emp is not None:
            df.loc[df["Id"] == uid, ['Name', 'Department']] = [emp.name, emp.department]
            df.to_csv(RECORDS_CSV_PATH, index=False, encoding='utf-8-sig')
    if invalid:
        session['manage_invalid'] = invalid
        if photo_error:
            session['photo_error'] = photo_error
        return _manage_redirect(qs, edit_uid=uid)
    return _manage_redirect(qs, edit_uid=None)
@app.route('/accounts/delete/<string:uid>')
@login_required
@require_role('admin')
def account_delete(uid):
    next_url = request.args.get('next')
    if uid == current_user.id:
        return redirect(next_url or '/manage?active_tab=accounts')
    u = users.query.filter_by(id=uid).first()
    target_tab = u.role if u is not None and u.role in ('teacher', 'student') else 'accounts'
    if u is not None:
        db.session.delete(u)
    emp = employee.query.filter_by(id=uid).first()
    if emp is not None:
        db.session.delete(emp)
    db.session.commit()
    _log_edit(current_user.id, uid, "xoá tài khoản")
    try:
        fname = os.path.join(path, uid + '.jpg')
        if os.path.exists(fname):
            os.remove(fname)
    except Exception as e:
        print("[account_delete] Lỗi xoá ảnh:", e)
    return redirect(next_url or f'/manage?active_tab={target_tab}')
@app.route('/accounts/export')
@login_required
@require_role('admin')
def accounts_export():
    all_users = users.query.order_by(users.role, users.name).all()
    emp_map = {e.id: e for e in employee.query.all()}
    headers = ['Mã số', 'Tên đăng nhập', 'Họ tên', 'Vai trò', 'Trạng thái',
               'Email', 'Lớp/Môn học', 'Nơi công tác', 'Chức vụ']
    rows = [[u.id, u.username, u.name, ROLE_LABELS.get(u.role, u.role),
            ACCOUNT_STATUS_LABELS.get(u.status, u.status), u.mail,
            emp_map.get(u.id).department if emp_map.get(u.id) else '',
            u.workplace or '', u.position or ''] for u in all_users]
    return _xlsx_response('DANH SÁCH TÀI KHOẢN', headers, rows, 'danh_sach_tai_khoan.xlsx')
@app.route('/accounts/history')
@login_required
@require_role('admin')
def accounts_history():
    logs = EditHistory.query.order_by(EditHistory.timestamp.desc()).all()
    return render_template('editHistory.html', logs=logs)
@app.route('/accounts/history/export')
@login_required
@require_role('admin')
def accounts_history_export():
    logs = EditHistory.query.order_by(EditHistory.timestamp.desc()).all()
    headers = ['Thời gian', 'Quản trị viên', 'Tài khoản bị sửa', 'Nội dung thay đổi']
    rows = [[log.timestamp.strftime('%d-%m-%Y %H:%M:%S'), log.admin_id, log.target_id, log.action] for log in logs]
    return _xlsx_response('LỊCH SỬ CHỈNH SỬA', headers, rows, 'lich_su_chinh_sua.xlsx')
@app.route('/profile')
@login_required
def profile():
    if current_user.role == 'admin':
        return redirect('/add')
    emp = employee.query.filter_by(id=current_user.id).first()
    return render_template('profilePage.html', emp=emp, user=current_user)
@app.route('/classes', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def classes_page():
    if request.method == 'POST':
        action = request.form.get('action')
        target_tab = 'classes'
        if action == 'add_class':
            name = request.form.get('class_name', '').strip()
            grade = request.form.get('class_grade', '').strip()
            if not grade:
                m = re.match(r'^\s*(\d+)', name)
                grade = m.group(1) if m else ''
            if name and not Class_.query.filter_by(name=name).first():
                db.session.add(Class_(name=name, grade=grade))
                db.session.commit()
        elif action == 'delete_class':
            c = Class_.query.filter_by(id=request.form.get('class_id')).first()
            if c:
                db.session.delete(c)
                db.session.commit()
        elif action == 'add_subject':
            target_tab = 'subjects'
            name = request.form.get('subject_name', '').strip()
            if name and not Subject.query.filter_by(name=name).first():
                db.session.add(Subject(name=name))
                db.session.commit()
        elif action == 'delete_subject':
            target_tab = 'subjects'
            s = Subject.query.filter_by(id=request.form.get('subject_id')).first()
            if s:
                db.session.delete(s)
                db.session.commit()
        elif action == 'assign':
            target_tab = 'assignments'
            teacher_id = request.form.get('teacher_id', '').strip()
            class_name = request.form.get('class_name', '').strip()
            if teacher_id and class_name and not TeacherAssignment.query.filter_by(
                    teacher_id=teacher_id, class_name=class_name).first():
                db.session.add(TeacherAssignment(teacher_id=teacher_id, class_name=class_name))
                db.session.commit()
        elif action == 'unassign':
            target_tab = 'assignments'
            a = TeacherAssignment.query.filter_by(id=request.form.get('assignment_id')).first()
            if a:
                db.session.delete(a)
                db.session.commit()
        return redirect(f'/manage?active_tab={target_tab}')
    return redirect('/manage?active_tab=classes')
@app.route('/assignments', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def assignments_page():
    return redirect('/manage?active_tab=assignments')
WEEKDAY_CODES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
WEEKDAY_LABELS_VI = {
    'Mon': 'Thứ 2', 'Tue': 'Thứ 3', 'Wed': 'Thứ 4', 'Thu': 'Thứ 5',
    'Fri': 'Thứ 6', 'Sat': 'Thứ 7', 'Sun': 'Chủ nhật',
}
def _reports_redirect(qs, **overrides):
    params = {k: v[0] for k, v in parse_qs(qs or '').items()}
    for k, v in overrides.items():
        if v is None:
            params.pop(k, None)
        else:
            params[k] = v
    params = {k: v for k, v in params.items() if v not in (None, '')}
    return redirect(url_for('reports') + (('?' + urlencode(params)) if params else ''))
@app.route("/sessions", methods=['GET', 'POST'])
@login_required
@require_role('admin')
def sessions():
    qs = request.form.get('qs', '') if request.method == 'POST' else request.query_string.decode('utf-8')
    if request.method != 'POST':
        return _reports_redirect(qs)
    name       = request.form.get('name', '').strip()
    days       = request.form.getlist('days')
    open_time  = request.form.get('open_time', '').strip()
    close_time = request.form.get('close_time', '').strip()
    if not name or not days or not open_time or not close_time:
        error = "Vui lòng nhập đủ tên buổi, chọn ít nhất 1 ngày, và giờ mở/đóng cổng."
        return _reports_redirect(qs, active_tab='sheet', session_error=error)
    if open_time >= close_time:
        error = "Giờ mở cổng phải TRƯỚC giờ đóng cổng."
        return _reports_redirect(qs, active_tab='sheet', session_error=error)
    try:
        new_session = AttendanceSession(
            name=name,
            days_of_week=",".join(days),
            open_time=open_time,
            close_time=close_time,
        )
        db.session.add(new_session)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        error = f"Lỗi khi lưu buổi điểm danh: {e}"
        return _reports_redirect(qs, active_tab='sheet', session_error=error)
    return _reports_redirect(qs, active_tab='sessionreport', session_id=str(new_session.id), session_error=None)
@app.route("/sessions/delete/<int:session_id>")
@login_required
def delete_session(session_id):
    qs = request.args.get('qs', '')
    s = AttendanceSession.query.get(session_id)
    if s:
        db.session.delete(s)
        db.session.commit()
    params = parse_qs(qs)
    overrides = {'active_tab': 'sheet'}
    if params.get('session_id', [''])[0] == str(session_id):
        overrides['session_id'] = None
    return _reports_redirect(qs, **overrides)
def _compute_day_report(s, picked_date, report_role='student', f_id='', f_dept='', allowed_ids=None):
    weekday_code = WEEKDAY_CODES[picked_date.weekday()]
    date_str = picked_date.strftime('%Y-%m-%d')
    date_display = picked_date.strftime('%d-%m-%Y')
    applies_today = weekday_code in s.day_list()
    emp_query = employee.query.filter_by(role=report_role)
    if allowed_ids is not None:
        if not allowed_ids:
            emp_query = emp_query.filter(employee.id == None)
        else:
            emp_query = emp_query.filter(employee.id.in_(allowed_ids))
    if f_id:
        like = f'%{f_id}%'
        emp_query = emp_query.filter(db.or_(employee.id.ilike(like), employee.name.ilike(like)))
    if f_dept:
        emp_query = emp_query.filter(employee.department == f_dept)
    all_employees = emp_query.order_by(employee.department, employee.name).all()
    ensure_records_csv()
    times_by_id = {}
    with open(RECORDS_CSV_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get('Date') != date_display:
                continue
            eid = row.get('Id')
            t   = row.get('Time', '')
            if eid and t:
                times_by_id.setdefault(eid, []).append(t)
    def pick_best_time(times):
        open_dt  = datetime.strptime(s.open_time, '%H:%M')
        close_dt = datetime.strptime(s.close_time, '%H:%M')
        tol = timedelta(seconds=60)
        early_window_seconds = 3600
        parsed = [(t, datetime.strptime(t, '%H:%M:%S')) for t in times]
        in_window = [p for p in parsed if open_dt - tol <= p[1] <= close_dt + tol]
        if in_window:
            best = min(in_window, key=lambda p: (p[1] - open_dt).total_seconds())
            return best[0], 'on_time'
        candidates = []
        for t, dt in parsed:
            if dt < open_dt - tol:
                dist = (open_dt - dt).total_seconds()
                if dist <= early_window_seconds:
                    candidates.append((t, dist, 'early'))
            elif dt > close_dt + tol:
                dist = (dt - close_dt).total_seconds()
                candidates.append((t, dist, 'late'))
        if not candidates:
            return None, None
        best = min(candidates, key=lambda c: c[1])
        return best[0], best[2]
    on_time, early, late, absent = [], [], [], []
    for emp in all_employees:
        times = times_by_id.get(emp.id)
        if not times:
            absent.append(emp)
            continue
        best_time, status = pick_best_time(times)
        if status == 'on_time':
            on_time.append((emp, best_time))
        elif status == 'early':
            early.append((emp, best_time))
        elif status == 'late':
            late.append((emp, best_time))
        else:
            absent.append(emp)
    return dict(
        date_str=date_str, date_display=date_display,
        weekday_code=weekday_code, weekday_label=WEEKDAY_LABELS_VI[weekday_code],
        applies_today=applies_today,
        on_time=on_time, early=early, late=late, absent=absent,
    )
def _compute_session_report_context(session_id, from_str=None, to_str=None, report_role='student', f_id='', f_dept='', allowed_ids=None):
    s = AttendanceSession.query.get(session_id)
    if s is None:
        return None
    today = datetime.now()
    try:
        to_date = datetime.strptime(to_str, '%Y-%m-%d') if to_str else today
    except ValueError:
        to_date = today
    try:
        from_date = datetime.strptime(from_str, '%Y-%m-%d') if from_str else to_date - timedelta(days=6)
    except ValueError:
        from_date = to_date - timedelta(days=6)
    if from_date > to_date:
        from_date, to_date = to_date, from_date
    applicable_days = s.day_list()
    dates = []
    cur = to_date
    while cur >= from_date:
        if not applicable_days or WEEKDAY_CODES[cur.weekday()] in applicable_days:
            dates.append(cur.strftime('%Y-%m-%d'))
        cur -= timedelta(days=1)
    report_days = [_compute_day_report(s, datetime.strptime(d, '%Y-%m-%d'), report_role, f_id, f_dept, allowed_ids) for d in dates]
    return dict(
        report_session=s,
        report_from=from_date.strftime('%Y-%m-%d'), report_to=to_date.strftime('%Y-%m-%d'),
        report_applicable_days=applicable_days, report_days=report_days,
    )
@app.route("/downloadAll")
@login_required
@require_role('admin')
def downloadAll():
    ensure_records_csv()
    return send_file(RECORDS_CSV_PATH, as_attachment=True)
@app.route("/downloadToday")
@login_required
@require_role('admin')
def downloadToday():
    ensure_records_csv()
    df = _read_records_df()
    df = df[df['Date'] == datetime.now().strftime("%d-%m-%Y")]
    df.to_csv(TODAY_CSV_PATH, index=False, encoding='utf-8-sig')
    return send_file(TODAY_CSV_PATH, as_attachment=True)
@app.route("/resetToday")
@login_required
@require_role('admin')
def resetToday():
    ensure_records_csv()
    df = _read_records_df()
    df = df[df['Date'] != datetime.now().strftime("%d-%m-%Y")]
    df.to_csv(RECORDS_CSV_PATH, index=False, encoding='utf-8-sig')
    return _reports_redirect(request.args.get('qs', ''), active_tab='sheet')
def _classify_record_status(row):
    try:
        d = datetime.strptime(row.get('Date', ''), '%d-%m-%Y')
        t = datetime.strptime(row.get('Time', ''), '%H:%M:%S')
    except (ValueError, TypeError):
        return 'Có mặt'
    weekday_code = WEEKDAY_CODES[d.weekday()]
    sessions = AttendanceSession.query.all()
    tol = timedelta(seconds=60)
    early_window_seconds = 3600
    for s in sessions:
        if weekday_code not in s.day_list():
            continue
        open_dt  = datetime.strptime(s.open_time, '%H:%M')
        close_dt = datetime.strptime(s.close_time, '%H:%M')
        check_dt = t.replace(year=open_dt.year, month=open_dt.month, day=open_dt.day)
        if open_dt - tol <= check_dt <= close_dt + tol:
            return 'Có mặt'
        if check_dt < open_dt - tol:
            if (open_dt - check_dt).total_seconds() <= early_window_seconds:
                return 'Đến sớm'
        elif check_dt > close_dt + tol:
            return 'Đến trễ'
    return 'Có mặt'
@app.route("/records")
@login_required
def records():
    ensure_records_csv()
    df = _read_records_df()
    rows = df.to_dict('records')
    for row in rows:
        row['Trạng thái'] = _classify_record_status(row)
    allowed = _allowed_employee_ids_for(current_user)
    if allowed is not None:
        rows = [r for r in rows if r.get('Id') in allowed]
    return render_template('RecordsPage.html', allrows=rows, fieldnames=RECORDS_FIELDS, len=len)
def _compute_session_stats_context(report_session, report_days):
    empty = dict(
        sess_totals=dict(on_time=0, early=0, late=0, absent=0), sess_total_all=0, sess_rate=0.0,
        sess_status_rows=[], sess_dept_rows=[], sess_daily_rows=[], sess_emp_rows=[],
    )
    if not report_session or not report_days:
        return empty
    totals = {'on_time': 0, 'early': 0, 'late': 0, 'absent': 0}
    dept_totals = {}
    emp_totals = {}
    daily_rows = []
    for day in sorted(report_days, key=lambda d: d['date_str']):
        counts = {'on_time': len(day['on_time']), 'early': len(day['early']),
                  'late': len(day['late']), 'absent': len(day['absent'])}
        for k in totals:
            totals[k] += counts[k]
        daily_rows.append({'date': day['date_display'], 'on_time': counts['on_time'],
                           'early': counts['early'], 'late': counts['late'],
                           'absent': counts['absent'], 'total': sum(counts.values())})
        for emp, t in day['on_time']:
            dept_totals.setdefault(emp.department, {'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['on_time'] += 1
            emp_totals.setdefault(emp.id, {'id': emp.id, 'name': emp.name, 'department': emp.department,
                                           'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['on_time'] += 1
        for emp, t in day['early']:
            dept_totals.setdefault(emp.department, {'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['early'] += 1
            emp_totals.setdefault(emp.id, {'id': emp.id, 'name': emp.name, 'department': emp.department,
                                           'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['early'] += 1
        for emp, t in day['late']:
            dept_totals.setdefault(emp.department, {'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['late'] += 1
            emp_totals.setdefault(emp.id, {'id': emp.id, 'name': emp.name, 'department': emp.department,
                                           'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['late'] += 1
        for emp in day['absent']:
            dept_totals.setdefault(emp.department, {'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['absent'] += 1
            emp_totals.setdefault(emp.id, {'id': emp.id, 'name': emp.name, 'department': emp.department,
                                           'on_time': 0, 'early': 0, 'late': 0, 'absent': 0})['absent'] += 1
    total_all = sum(totals.values())
    rate = round(totals['on_time'] / total_all * 100, 1) if total_all > 0 else 0.0
    status_labels = {'on_time': 'Có mặt', 'early': 'Đến sớm', 'late': 'Đến trễ', 'absent': 'Vắng mặt'}
    status_rows = [{
        'label': status_labels[k], 'count': totals[k],
        'percent': round(totals[k] / total_all * 100, 1) if total_all > 0 else 0.0,
    } for k in ('on_time', 'early', 'late', 'absent')]
    dept_rows = []
    for d, v in sorted(dept_totals.items()):
        d_total = sum(v.values())
        dept_rows.append({
            'department': d, 'on_time': v['on_time'], 'early': v['early'], 'late': v['late'],
            'absent': v['absent'], 'total': d_total,
            'rate': round(v['on_time'] / d_total * 100, 1) if d_total > 0 else 0.0,
        })
    emp_rows = sorted(emp_totals.values(), key=lambda r: r['name'])
    return dict(
        sess_totals=totals, sess_total_all=total_all, sess_rate=rate,
        sess_status_rows=status_rows, sess_dept_rows=dept_rows, sess_daily_rows=daily_rows,
        sess_emp_rows=emp_rows,
    )
def _row_date(row):
    try:
        return datetime.strptime(row.get('Date', ''), '%d-%m-%Y')
    except ValueError:
        return datetime.min
def _filter_records(rows, f_id, f_dept, f_status, f_from, f_to):
    if f_id:
        rows = [r for r in rows if f_id.lower() in r.get('Id', '').lower()
                or f_id.lower() in r.get('Name', '').lower()]
    if f_dept:
        rows = [r for r in rows if f_dept.lower() in r.get('Department', '').lower()]
    if f_status:
        rows = [r for r in rows if r.get('Trạng thái', '') == f_status]
    if f_from:
        try:
            d_from = datetime.strptime(f_from, '%Y-%m-%d')
            rows = [r for r in rows if _row_date(r) >= d_from]
        except ValueError:
            pass
    if f_to:
        try:
            d_to = datetime.strptime(f_to, '%Y-%m-%d')
            rows = [r for r in rows if _row_date(r) <= d_to]
        except ValueError:
            pass
    rows.sort(key=lambda r: (_row_date(r), r.get('Time', '')), reverse=True)
    return rows
HEALTH_STATUS_META = {
    'good':    {'label': 'Tốt',              'color': 'success',   'icon': 'fa-check-circle'},
    'normal':  {'label': 'Bình thường',      'color': 'info',      'icon': 'fa-circle'},
    'warning': {'label': 'Cần theo dõi',     'color': 'warning',   'icon': 'fa-exclamation-triangle'},
    'danger':  {'label': 'Nguy hiểm',        'color': 'danger',    'icon': 'fa-skull-crossbones'},
    'unknown': {'label': 'Chưa có dữ liệu',  'color': 'secondary', 'icon': 'fa-question-circle'},
}
_HEALTH_RANK = {'unknown': 0, 'good': 1, 'normal': 2, 'warning': 3, 'danger': 4}
def _classify_heart_rate(v):
    if v is None:
        return 'unknown'
    if 60 <= v <= 100:
        return 'good'
    if 50 <= v < 60 or 100 < v <= 110:
        return 'normal'
    if 40 <= v < 50 or 110 < v <= 130:
        return 'warning'
    return 'danger'
def _classify_spo2(v):
    if v is None:
        return 'unknown'
    if v >= 95:
        return 'good'
    if v >= 90:
        return 'normal'
    if v >= 85:
        return 'warning'
    return 'danger'
def _classify_temperature(v):
    if v is None:
        return 'unknown'
    if 36.1 <= v <= 37.2:
        return 'good'
    if 37.2 < v <= 37.8 or 35.5 <= v < 36.1:
        return 'normal'
    if 37.8 < v <= 38.5 or 35.0 <= v < 35.5:
        return 'warning'
    return 'danger'
def _overall_health_status(*statuses):
    return max(statuses, key=lambda st: _HEALTH_RANK[st])
def _compute_health_context(f_health_status='', report_role='student', f_id='', f_dept='', allowed_ids=None):
    rows = []
    emp_query = employee.query.filter_by(role=report_role)
    if allowed_ids is not None:
        if not allowed_ids:
            emp_query = emp_query.filter(employee.id == None)
        else:
            emp_query = emp_query.filter(employee.id.in_(allowed_ids))
    if f_id:
        like = f'%{f_id}%'
        emp_query = emp_query.filter(db.or_(employee.id.ilike(like), employee.name.ilike(like)))
    if f_dept:
        emp_query = emp_query.filter(employee.department == f_dept)
    for emp in emp_query.order_by(employee.department, employee.name).all():
        heart_rate = None
        spo2 = None
        temperature = None
        hr_status = _classify_heart_rate(heart_rate)
        spo2_status = _classify_spo2(spo2)
        temp_status = _classify_temperature(temperature)
        overall_status = _overall_health_status(hr_status, spo2_status, temp_status)
        rows.append(dict(
            employee=emp,
            heart_rate=heart_rate, spo2=spo2, temperature=temperature,
            hr_status=hr_status, spo2_status=spo2_status, temp_status=temp_status,
            overall_status=overall_status,
        ))
    if f_health_status:
        rows = [r for r in rows if r['overall_status'] == f_health_status]
    return dict(health_rows=rows, health_status_meta=HEALTH_STATUS_META, f_health_status=f_health_status,
                report_role=report_role, role_label=ROLE_LABELS[report_role], role_dept_label=ROLE_DEPT_LABEL[report_role])
def _render_print_report(title, meta_lines, headers, rows, summary_title=None, summary_headers=None, summary_rows=None):
    meta_html = ''.join(f'<p>{m}</p>' for m in meta_lines)
    thead = ''.join(f'<th>{h}</th>' for h in headers)
    if rows:
        tbody = ''.join('<tr>' + ''.join(f'<td>{c}</td>' for c in r) + '</tr>' for r in rows)
    else:
        tbody = f'<tr><td colspan="{len(headers)}" style="text-align:center;">Không có dữ liệu</td></tr>'
    summary_html = ''
    if summary_headers:
        s_thead = ''.join(f'<th>{h}</th>' for h in summary_headers)
        if summary_rows:
            s_tbody = ''.join('<tr>' + ''.join(f'<td>{c}</td>' for c in r) + '</tr>' for r in summary_rows)
        else:
            s_tbody = f'<tr><td colspan="{len(summary_headers)}" style="text-align:center;">Không có dữ liệu</td></tr>'
        summary_html = (f'<h2 class="summary-title">{summary_title or "Thống kê"}</h2>'
                         f'<table><thead><tr>{s_thead}</tr></thead><tbody>{s_tbody}</tbody></table>')
    return f"""<!doctype html>
<html lang="vi"><head><meta charset="utf-8">
<title>{title}</title>
<style>
body {{ font-family: 'Times New Roman', Times, serif; color: #000; padding: 24px; }}
h1 {{ text-align: center; font-size: 22px; margin-bottom: 4px; }}
.summary-title {{ text-align: center; font-size: 18px; margin-top: 32px; }}
.meta p {{ margin: 2px 0; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 16px; }}
th, td {{ border: 1px solid #000; padding: 6px 8px; font-size: 14px; }}
th {{ background: #ddd; }}
@media print {{ body {{ padding: 0; }} }}
</style></head>
<body>
<h1>{title}</h1>
<div class="meta">{meta_html}</div>
<table><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table>
{summary_html}
<script>
window.onload = function() {{
window.print();
}};
</script>
</body></html>"""
@app.route("/reports/students")
@login_required
def reports_students():
    return redirect(_reports_redirect(request.query_string.decode(), report_role='student'))
@app.route("/reports/teachers")
@login_required
def reports_teachers():
    return redirect(_reports_redirect(request.query_string.decode(), report_role='teacher'))
@app.route("/reports")
@login_required
def reports():
    ensure_records_csv()
    rows = []
    with open(RECORDS_CSV_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    f_id     = request.args.get('f_id', '').strip()
    f_dept   = request.args.get('f_dept', '').strip()
    f_status = request.args.get('f_status', '').strip()
    f_from   = request.args.get('f_from', '').strip()
    f_to     = request.args.get('f_to', '').strip()
    report_role = request.args.get('report_role', 'student').strip()
    if report_role not in ROLE_LABELS:
        report_role = 'student'
    rows = _filter_records(rows, f_id, f_dept, f_status, f_from, f_to)
    role_ids = {e.id for e in employee.query.filter_by(role=report_role).all()}
    rows = [r for r in rows if r.get('Id') in role_ids]
    allowed = _allowed_employee_ids_for(current_user)
    if allowed is not None:
        rows = [r for r in rows if r.get('Id') in allowed]
    statuses = sorted({r.get('Trạng thái', '') for r in rows if r.get('Trạng thái')})
    all_sessions = AttendanceSession.query.all()
    f_health_status = request.args.get('f_health_status', '').strip()
    health_ctx = _compute_health_context(f_health_status, report_role, f_id, f_dept, allowed_ids=allowed)
    sel_session_id = request.args.get('session_id', type=int)
    sel_report_from = request.args.get('report_from', '').strip()
    sel_report_to = request.args.get('report_to', '').strip()
    active_tab = request.args.get('active_tab', '').strip()
    if not active_tab:
        active_tab = 'sessionreport' if sel_session_id else 'sheet'
    report_ctx = {}
    if sel_session_id:
        report_ctx = _compute_session_report_context(sel_session_id, sel_report_from, sel_report_to, report_role, f_id, f_dept, allowed) or {}
    stats_ctx = _compute_session_stats_context(report_ctx.get('report_session'), report_ctx.get('report_days'))
    base_params = request.args.to_dict(flat=True)
    base_params.pop('report_role', None)
    qs_role_student = urlencode({**base_params, 'report_role': 'student'})
    qs_role_teacher = urlencode({**base_params, 'report_role': 'teacher'})
    resp = render_template(
        'Reportspage.html',
        allrows=rows, statuses=statuses,
        report_generated_at=datetime.now().strftime('%d-%m-%Y %H:%M'),
        f_id=f_id, f_dept=f_dept, f_status=f_status, f_from=f_from, f_to=f_to,
        all_sessions=all_sessions,
        weekday_codes=WEEKDAY_CODES, weekday_labels=WEEKDAY_LABELS_VI,
        sel_session_id=sel_session_id, active_tab=active_tab,
        current_qs=request.query_string.decode('utf-8'),
        current_qs_q=quote(request.query_string.decode('utf-8')),
        session_error=request.args.get('session_error', ''),
        qs_role_student=qs_role_student, qs_role_teacher=qs_role_teacher,
        **stats_ctx,
        **health_ctx,
        **report_ctx,
    )
    response = app.make_response(resp)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return response
@app.route("/reports/export")
@login_required
def reports_export():
    export_type = request.args.get('type', 'sheet').strip()
    report_role = request.args.get('report_role', 'student').strip()
    if report_role not in ROLE_LABELS:
        report_role = 'student'
    dept_label = ROLE_DEPT_LABEL[report_role]
    allowed = _allowed_employee_ids_for(current_user)
    if export_type == 'health':
        f_health_status = request.args.get('f_health_status', '').strip()
        f_id = request.args.get('f_id', '').strip()
        f_dept = request.args.get('f_dept', '').strip()
        health_ctx = _compute_health_context(f_health_status, report_role, f_id, f_dept, allowed_ids=allowed)
        headers = ['Id', 'Tên', dept_label, 'Nhịp tim', 'SpO₂', 'Nhiệt độ', 'Trạng thái']
        table_rows = [[
            row['employee'].id, row['employee'].name, row['employee'].department,
            f"{row['heart_rate']} bpm" if row['heart_rate'] is not None else '--',
            f"{row['spo2']}%" if row['spo2'] is not None else '--',
            f"{row['temperature']}°C" if row['temperature'] is not None else '--',
            HEALTH_STATUS_META[row['overall_status']]['label'],
        ] for row in health_ctx['health_rows']]
        health_total = len(health_ctx['health_rows'])
        status_counts = {k: 0 for k in HEALTH_STATUS_META}
        for row in health_ctx['health_rows']:
            status_counts[row['overall_status']] += 1
        summary_headers = ['Trạng thái', 'Số lượng', 'Tỉ lệ']
        summary_rows = [[
            HEALTH_STATUS_META[k]['label'], status_counts[k],
            f"{round(status_counts[k] / health_total * 100, 1)}%" if health_total > 0 else '0%',
        ] for k in HEALTH_STATUS_META]
        summary_rows.append(['Tổng', health_total, '100%' if health_total > 0 else '0%'])
        meta_lines = [
            f"Loại báo cáo: Sức khỏe {ROLE_LABELS[report_role]}",
            f"Ngày lập: {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            f"Người lập: {current_user.username if current_user.is_authenticated else ''}",
            f"Tổng số: {health_total}",
        ]
        html = _render_print_report(f"BÁO CÁO SỨC KHỎE - {ROLE_LABELS[report_role].upper()}", meta_lines, headers, table_rows,
                                    'Thống kê trạng thái sức khỏe', summary_headers, summary_rows)
        return Response(html, mimetype='text/html; charset=utf-8')
    if export_type == 'sessionreport':
        session_id = request.args.get('session_id', type=int)
        report_from = request.args.get('report_from', '').strip()
        report_to = request.args.get('report_to', '').strip()
        f_id = request.args.get('f_id', '').strip()
        f_dept = request.args.get('f_dept', '').strip()
        ctx = _compute_session_report_context(session_id, report_from, report_to, report_role, f_id, f_dept, allowed) if session_id else None
        headers = ['Ngày', 'Thứ', 'Id', 'Tên', dept_label, 'Trạng thái', 'Giờ vào']
        table_rows = []
        if ctx:
            for day in ctx['report_days']:
                for emp, t in day['on_time']:
                    table_rows.append([day['date_display'], day['weekday_label'], emp.id, emp.name, emp.department, 'Có mặt', t])
                for emp, t in day['early']:
                    table_rows.append([day['date_display'], day['weekday_label'], emp.id, emp.name, emp.department, 'Đến sớm', t])
                for emp, t in day['late']:
                    table_rows.append([day['date_display'], day['weekday_label'], emp.id, emp.name, emp.department, 'Đến trễ', t])
                for emp in day['absent']:
                    table_rows.append([day['date_display'], day['weekday_label'], emp.id, emp.name, emp.department, 'Vắng mặt', ''])
        stats_ctx = _compute_session_stats_context(ctx.get('report_session') if ctx else None, ctx.get('report_days') if ctx else None)
        summary_headers = ['Trạng thái', 'Số lượt', 'Tỉ lệ']
        summary_rows = [[r['label'], r['count'], f"{r['percent']}%"] for r in stats_ctx['sess_status_rows']]
        summary_rows.append(['Tổng', stats_ctx['sess_total_all'], '100%' if stats_ctx['sess_total_all'] > 0 else '0%'])
        meta_lines = [
            f"Loại báo cáo: {ROLE_LABELS[report_role]}",
            f"Buổi điểm danh: {ctx['report_session'].name if ctx else ''}",
            f"Khoảng thời gian: {ctx['report_from'] if ctx else ''} - {ctx['report_to'] if ctx else ''}",
            f"Ngày lập: {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            f"Người lập: {current_user.username if current_user.is_authenticated else ''}",
        ]
        html = _render_print_report(f"BÁO CÁO BUỔI ĐIỂM DANH - {ROLE_LABELS[report_role].upper()}", meta_lines, headers, table_rows,
                                    'Thống kê trạng thái điểm danh', summary_headers, summary_rows)
        return Response(html, mimetype='text/html; charset=utf-8')
    ensure_records_csv()
    rows = []
    with open(RECORDS_CSV_PATH, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    rows = _filter_records(
        rows,
        request.args.get('f_id', '').strip(), request.args.get('f_dept', '').strip(),
        request.args.get('f_status', '').strip(), request.args.get('f_from', '').strip(),
        request.args.get('f_to', '').strip(),
    )
    role_ids = {e.id for e in employee.query.filter_by(role=report_role).all()}
    rows = [r for r in rows if r.get('Id') in role_ids]
    if allowed is not None:
        rows = [r for r in rows if r.get('Id') in allowed]
    headers = ['Id', 'Tên', dept_label, 'Giờ', 'Ngày', 'Trạng thái']
    table_rows = [[r.get('Id',''), r.get('Name',''), r.get('Department',''), r.get('Time',''), r.get('Date',''), r.get('Trạng thái','')] for r in rows]
    sheet_total = len(rows)
    sheet_status_counts = {}
    for r in rows:
        st = r.get('Trạng thái', '')
        sheet_status_counts[st] = sheet_status_counts.get(st, 0) + 1
    summary_headers = ['Trạng thái', 'Số lượng', 'Tỉ lệ']
    summary_rows = [[
        st, cnt, f"{round(cnt / sheet_total * 100, 1)}%" if sheet_total > 0 else '0%',
    ] for st, cnt in sorted(sheet_status_counts.items())]
    summary_rows.append(['Tổng', sheet_total, '100%' if sheet_total > 0 else '0%'])
    meta_lines = [
        f"Loại báo cáo: {ROLE_LABELS[report_role]}",
        f"Ngày lập: {datetime.now().strftime('%d-%m-%Y %H:%M')}",
        f"Người lập: {current_user.username if current_user.is_authenticated else ''}",
        f"Tổng số bản ghi: {sheet_total}",
    ]
    html = _render_print_report(f"BÁO CÁO ĐIỂM DANH - {ROLE_LABELS[report_role].upper()}", meta_lines, headers, table_rows,
                                'Thống kê trạng thái', summary_headers, summary_rows)
    return Response(html, mimetype='text/html; charset=utf-8')
@app.route('/reset_request', methods=['GET', 'POST'])
def reset_request():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        id_      = request.form.get('id', '').strip()
        name     = request.form.get('name', '').strip()
        email    = request.form.get('mail', '').strip()
        user = users.query.filter_by(username=username, id=id_, name=name, mail=email).first()
        if user is None:
            return render_template('resetRequest.html', incorrect=True,
                                   msg="Thông tin không khớp. Vui lòng kiểm tra lại tên đăng nhập, "
                                       "mã số, họ tên và email.")
        if user.status == 'disabled':
            return render_template('resetRequest.html', incorrect=True,
                                   msg="Tài khoản đã bị vô hiệu hoá. Vui lòng liên hệ quản trị viên.")
        otp = randint(100000, 999999)
        ok, err = _sendResetMail(email, otp)
        if not ok:
            if err and err.startswith("smtp_auth:"):
                friendly_msg = err.split(":", 1)[1]
            elif err and err.startswith("smtp_other:"):
                friendly_msg = err.split(":", 1)[1]
            else:
                friendly_msg = ("Không gửi được email xác nhận do lỗi không "
                                "lường trước. Vui lòng liên hệ quản trị viên "
                                "hệ thống để được hỗ trợ.")
            return render_template('resetRequest.html', incorrect=True,
                                   msg=friendly_msg)
        session['id']  = user.id
        session['otp'] = otp
        return render_template('OTP.html')
    return render_template('resetRequest.html')
def _sendResetMail(mail, otp):
    try:
        msg = Message('Mã xác thực OTP - Hệ thống điểm danh bằng nhận diện khuôn mặt', recipients=[mail],
                      sender=app.config['MAIL_DEFAULT_SENDER'])
        msg.body = (
            "Xin chào,\n\n"
            "Bạn vừa yêu cầu mã xác thực cho tài khoản trên Hệ thống điểm danh bằng nhận diện khuôn mặt.\n\n"
            f"Mã OTP của bạn là: {otp}\n\n"
            "Vui lòng nhập mã này để hoàn tất xác thực. Nếu bạn không thực hiện yêu cầu này, "
            "vui lòng bỏ qua email và không chia sẻ mã cho bất kỳ ai.\n\n"
            "Trân trọng,\n"
            "Hệ thống điểm danh bằng nhận diện khuôn mặt"
        )
        mail_.send(msg)
        return True, None
    except smtplib.SMTPAuthenticationError as e:
        print(f"[reset_request] LỖI XÁC THỰC SMTP (535): {e}")
        traceback.print_exc()
        return False, (
            "smtp_auth:Không gửi được email — tài khoản gửi mail bị Gmail từ "
            "chối đăng nhập (sai mật khẩu hoặc chưa dùng Mật khẩu ứng dụng). "
            "Vui lòng liên hệ quản trị viên hệ thống để kiểm tra lại cấu hình "
            "email (MAIL_USERNAME/MAIL_PASSWORD)."
        )
    except (smtplib.SMTPException, OSError) as e:
        print(f"[reset_request] LỖI SMTP/mạng khi gửi mail: {type(e).__name__}: {e}")
        traceback.print_exc()
        return False, (
            "smtp_other:Không thể kết nối tới máy chủ email lúc này. Vui lòng "
            "thử lại sau ít phút hoặc liên hệ quản trị viên hệ thống."
        )
    except Exception as e:
        print(f"[reset_request] LỖI không lường trước khi gửi email OTP: {type(e).__name__}: {e}")
        traceback.print_exc()
        return False, f"other:{e}"
@app.route('/verifyOTP', methods=['GET', 'POST'])
def verifyOTP():
    otp_raw = request.form.get('otp')
    if 'otp' not in session or otp_raw is None:
        return redirect('/login')
    try:
        otp2 = int(otp_raw.strip())
    except ValueError:
        return render_template('OTP.html', incorrect=True)
    if session['otp'] != otp2:
        return render_template('OTP.html', incorrect=True)
    session.pop('otp', None)
    if 'pending_admin' in session:
        data = session.pop('pending_admin')
        db.session.add(users(
            id=data['id'], name=data['name'], mail=data['mail'],
            username=data['username'], password=data['pass1'],
            role='admin', status='active',
            workplace=data['workplace'], position=data['position'],
        ))
        db.session.commit()
        return render_template('login.html', registered=True)
    if 'pending_employee' in session:
        data = session.pop('pending_employee')
        try:
            emp = employee(id=data['id'], name=data['name'], department=data['dept'],
                           email=data['mail'], role=data['role'])
            db.session.add(emp)
            db.session.add(users(
                id=data['id'], name=data['name'], mail=data['mail'],
                username=data['username'], password=data['pass1'],
                role=data['role'], status='pending',
            ))
            db.session.commit()
            tmp_photo_path = data.get('photo_tmp_path')
            if tmp_photo_path and os.path.exists(tmp_photo_path):
                os.replace(tmp_photo_path, os.path.join(path, data['id'] + '.jpg'))
        except Exception as e:
            print("[verifyOTP] Lỗi tạo tài khoản nhân sự:", e)
            db.session.rollback()
        return render_template('login.html', registered=True)
    if 'id' in session:
        return render_template('resetPassword.html')
    return redirect('/login')
@app.route('/resetPass', methods=['GET', 'POST'])
def resetPass():
    if 'id' not in session:
        return redirect('/reset_request')
    pw1 = request.form['pass1']
    pw2 = request.form['pass2']
    if pw1 != pw2:
        return render_template('resetPassword.html', incorrect=True)
    user = users.query.filter_by(id=session['id']).first()
    if user is None:
        return redirect('/reset_request')
    user.password = pw1
    db.session.commit()
    return render_template('login.html', reseted=True)
@app.route('/get')
def get_bot_response():
    userText = request.args.get('msg')
    return bot_responses.get(userText, "Sorry, Can't help with it :(")
@app.route('/helpBot')
def helpBot():
    global bot_responses
    with open(HELP_JSON_PATH, encoding='utf-8') as f:
        bot_responses = json.load(f)
    return render_template('chatBot.html', keys=[*bot_responses])
@app.errorhandler(Exception)
def handle_unexpected_error(e):
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    print(f"[GLOBAL ERROR] Lỗi không lường trước tại {request.path}: {e}")
    traceback.print_exc()
    return jsonify({
        "error": "Đã xảy ra lỗi không lường trước ở server.",
        "detail": f"{type(e).__name__}: {e}",
        "path": request.path,
    }), 500
def _ensure_employee_role_column():
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'employee' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('employee')}
    if 'role' not in existing_cols:
        with db.engine.connect() as conn:
            conn.execute(text(
                "ALTER TABLE employee ADD COLUMN role VARCHAR(20) DEFAULT 'student'"
            ))
            conn.execute(text(
                "UPDATE employee SET role = 'student' WHERE role IS NULL"
            ))
            conn.commit()
def _ensure_class_grade_column():
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'class_' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('class_')}
    if 'grade' not in existing_cols:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE class_ ADD COLUMN grade VARCHAR(10)"))
            conn.commit()
def _class_sort_key(c):
    m = re.match(r'^\s*(\d+)', c.grade or c.name or '')
    grade_num = int(m.group(1)) if m else 999
    return (grade_num, c.name or '')
def _sorted_classes():
    return sorted(Class_.query.all(), key=_class_sort_key)
def _ensure_users_extra_columns():
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'users' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('users')}
    additions = [
        ('role', "ALTER TABLE users ADD COLUMN role VARCHAR(20) DEFAULT 'admin'"),
        ('status', "ALTER TABLE users ADD COLUMN status VARCHAR(20) DEFAULT 'active'"),
        ('workplace', "ALTER TABLE users ADD COLUMN workplace VARCHAR(80)"),
        ('position', "ALTER TABLE users ADD COLUMN position VARCHAR(80)"),
    ]
    with db.engine.connect() as conn:
        for col_name, stmt in additions:
            if col_name not in existing_cols:
                conn.execute(text(stmt))
        conn.execute(text("UPDATE users SET role = 'admin' WHERE role IS NULL"))
        conn.execute(text("UPDATE users SET status = 'active' WHERE status IS NULL"))
        conn.commit()
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        _ensure_employee_role_column()
        _ensure_users_extra_columns()
        _ensure_class_grade_column()
    ensure_records_csv()
    app.run(debug=True, threaded=True, use_reloader=False)